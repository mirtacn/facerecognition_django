from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db import transaction
from django.contrib.auth import authenticate, login, get_user_model
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django import forms
from django.db.models import Sum, F, FloatField, ExpressionWrapper
from django.db.models.functions import Coalesce
from django.utils.translation import gettext_lazy as _
import math
from django.db.models import Sum, F, Case, When, Value
from django.db.models.functions import Coalesce
from django.utils.safestring import mark_safe
from datetime import datetime, date, timedelta 
from .forms import Step1Form, Step2Form, Step3Form
from .models import (
    Mahasiswa, FotoWajah, Kegiatan_PA, Jenjang_Pendidikan,
    Tahun_Ajaran, Dosen, Mahasiswa_Dosen, Pengajuan_Pendaftaran,
    Status_Pemenuhan_SKS, Semester, FotoWajah, Mahasiswa_Dosen, Presensi,Durasi,VerificationLog, Dosen, Prodi, Kelas,
)
from django.core.cache import cache
from .captcha_utils import get_or_create_captcha, verify_captcha
from .forms import FilterRekapPresensiForm
from django.db.models import Sum, F, Case, When, Value
from django.db.models.functions import Coalesce
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.utils import timezone
from django.utils import translation
import zipfile
import io
import os
import base64
from django.contrib.auth import get_user_model
from django.core.files.base import ContentFile
from django.db.models import Q, Count
from django.contrib.auth import logout
from django.shortcuts import redirect
import json
from django.utils.safestring import mark_safe
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import logging
from liveness_detection import process_frame_liveness, init_liveness_detection, reset_detection_state
import threading
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.core.mail import send_mail
from django.conf import settings

detection_locks = {}
detection_lock = threading.Lock()


def register_wizard(request, step=1):
    """Wizard registrasi mahasiswa dengan 3 langkah"""
    User = get_user_model()
    
    step1_data = request.session.get('step1_data', {})
    step2_data = request.session.get('step2_data', {})
    form = None

    # ==================== STEP 1: AKUN & IDENTITAS ====================
    if step == 1:
        form = Step1Form(request.POST or None, initial=step1_data)
        
        if request.method == 'POST' and form.is_valid():
            email = form.cleaned_data.get('email')
            nim = form.cleaned_data.get('nim')
            
            # ========== PERBAIKAN VALIDASI ==========
            # Cari user dengan email/nim yang sama
            email_exists = User.objects.filter(email=email).first()
            nim_exists = User.objects.filter(username=nim).first()  # Pakai username, bukan nrp
            
            email_conflict = False
            nim_conflict = False
            
            # Validasi EMAIL
            if email_exists:
                if email_exists.status_akun == 'aktif' and email_exists.is_active:
                    email_conflict = True
                    form.add_error('email', 'Email sudah terdaftar dan AKTIF. Gunakan email lain.')
                elif email_exists.status_akun == 'pending':
                    email_conflict = True
                    form.add_error('email', f'Email {email} masih dalam proses verifikasi. Silakan tunggu.')
                else:
                    # Status ditolak - izinkan daftar ulang
                    print(f"INFO: Email {email} ditolak, izinkan daftar ulang")
            
            # Validasi NIM (username)
            if nim_exists:
                if nim_exists.status_akun == 'aktif' and nim_exists.is_active:
                    nim_conflict = True
                    form.add_error('nim', 'NIM sudah terdaftar dan AKTIF. Gunakan NIM lain.')
                elif nim_exists.status_akun == 'pending':
                    nim_conflict = True
                    form.add_error('nim', f'NIM {nim} masih dalam proses verifikasi. Silakan tunggu.')
                else:
                    # Status ditolak - izinkan daftar ulang
                    print(f"INFO: NIM {nim} ditolak, izinkan daftar ulang")
            
            # Jika tidak ada konflik, lanjutkan
            if not email_conflict and not nim_conflict:
                # Simpan data ke session (lanjut step 2)
                step1_data_save = {
                    'nama_lengkap': form.cleaned_data.get('nama_lengkap'),
                    'nim': form.cleaned_data.get('nim'),
                    'email': form.cleaned_data.get('email'),
                    'password': form.cleaned_data.get('password'),
                }
                
                # Simpan jenjang
                jenjang_obj = form.cleaned_data.get('jenjang')
                if jenjang_obj:
                    step1_data_save['jenjang_id'] = jenjang_obj.id
                    step1_data_save['jenjang_nama'] = jenjang_obj.nama_jenjang
                
                # Simpan prodi
                prodi_obj = form.cleaned_data.get('prodi')
                if prodi_obj:
                    step1_data_save['prodi_id'] = prodi_obj.id
                    step1_data_save['prodi_nama'] = prodi_obj.nama_prodi
                
                # Simpan kelas
                kelas_obj = form.cleaned_data.get('kelas')
                if kelas_obj:
                    step1_data_save['kelas_id'] = kelas_obj.id
                    step1_data_save['kelas_nama'] = kelas_obj.nama_kelas
                
                request.session['step1_data'] = step1_data_save
                request.session.pop('step2_data', None)
                
                return redirect('register_step', step=2)
        
        progress = 33
        return render(request, 'mahasiswa/register_step1.html', {
            'form': form, 
            'step': step, 
            'progress': progress
        })
    
    # ==================== STEP 2: AKADEMIK ====================
    # ==================== STEP 2: AKADEMIK ====================
    elif step == 2:
        # Validasi step1 sudah diisi
        if not step1_data:
            messages.warning(request, 'Silakan lengkapi data identitas terlebih dahulu.')
            return redirect('register_step', step=1)
        
        # Ambil data dari step1
        jenjang_id = step1_data.get('jenjang_id')
        prodi_id = step1_data.get('prodi_id')
        kelas_id = step1_data.get('kelas_id')
        
        jenjang_obj = None
        prodi_obj = None
        
        if jenjang_id:
            jenjang_obj = Jenjang_Pendidikan.objects.get(id=jenjang_id)
        
        if prodi_id:
            prodi_obj = Prodi.objects.get(id=prodi_id)
        
        # === INI YANG DITAMBAHKAN - definisikan initial_data ===
        initial_data = {}
        if step2_data:
            # Ambil data dari session jika ada
            semester_id = step2_data.get('semester')
            if semester_id:
                try:
                    initial_data['semester'] = Semester.objects.get(id=semester_id)
                except Semester.DoesNotExist:
                    pass
            
            dosen1_id = step2_data.get('dosen_pembimbing1')
            if dosen1_id:
                try:
                    initial_data['dosen_pembimbing1'] = Dosen.objects.get(id=dosen1_id)
                except Dosen.DoesNotExist:
                    pass
            
            dosen2_id = step2_data.get('dosen_pembimbing2')
            if dosen2_id:
                try:
                    initial_data['dosen_pembimbing2'] = Dosen.objects.get(id=dosen2_id)
                except Dosen.DoesNotExist:
                    pass
            
            dosen3_id = step2_data.get('dosen_pembimbing3')
            if dosen3_id:
                try:
                    initial_data['dosen_pembimbing3'] = Dosen.objects.get(id=dosen3_id)
                except Dosen.DoesNotExist:
                    pass
            
            kegiatan_ids = step2_data.get('kegiatan_pa_diambil', [])
            if kegiatan_ids:
                initial_data['kegiatan_pa_diambil'] = Kegiatan_PA.objects.filter(id__in=kegiatan_ids)
        # === SAMPAI SINI ===
        
        # Buat form dengan jenjang_obj dan initial_data
        form = Step2Form(request.POST or None, initial=initial_data, jenjang_obj=jenjang_obj)

        if request.method == 'POST' and form.is_valid():
            data = form.cleaned_data
            
            # Ambil semester untuk mendapatkan namanya
            semester_obj = data.get('semester')
            
            # Ambil daftar kegiatan PA untuk summary
            kegiatan_list = data.get('kegiatan_pa_diambil', [])
            kegiatan_nama_list = [k.nama_kegiatan for k in kegiatan_list]
            
            save_data = {
                'jenjang_id': jenjang_id,
                'jenjang_nama': jenjang_obj.nama_jenjang if jenjang_obj else None,
                'prodi_id': prodi_id,
                'prodi_nama': step1_data.get('prodi_nama'),
                'kelas_id': kelas_id,
                'kelas_nama': step1_data.get('kelas_nama'),
                'semester': semester_obj.id if semester_obj else None,
                'semester_nama': semester_obj.nama_semester if semester_obj else None,
                'semester_nomor': semester_obj.nomor_semester if semester_obj else None,
                'dosen_pembimbing1': data['dosen_pembimbing1'].id if data['dosen_pembimbing1'] else None,
                'dosen_pembimbing2': data['dosen_pembimbing2'].id if data['dosen_pembimbing2'] else None,
                'dosen_pembimbing3': data['dosen_pembimbing3'].id if data['dosen_pembimbing3'] else None,
                'kegiatan_pa_diambil': [k.id for k in kegiatan_list],
                'kegiatan_pa_nama': kegiatan_nama_list,
            }
            request.session['step2_data'] = save_data
            return redirect('register_step', step=3)
        
        progress = 67
        return render(request, 'mahasiswa/register_step2.html', {
            'form': form,
            'step': step,
            'progress': progress,
            'jenjang_obj': jenjang_obj,
            'prodi_obj': prodi_obj,
            'dosen_list': Dosen.objects.all().order_by('nama_dosen'),  
        })

    # ==================== STEP 3: FOTO & FINALISASI ====================
    elif step == 3:
        if not step1_data or not step2_data:
            messages.warning(request, "Sesi kadaluarsa. Ulangi dari awal.")
            return redirect('register_step', step=1)

        # PERBAIKAN: Handle POST request untuk submit
        if request.method == 'POST':
            # Panggil fungsi submit_registration
            return submit_registration(request)
        
        form = Step3Form()
        progress = 100
        
        return render(request, 'mahasiswa/register_step3.html', {
            'form': form, 
            'step': step, 
            'progress': progress,
        })
        
@csrf_exempt
def submit_registration(request):
    print("="*50)
    print("SUBMIT REGISTRATION CALLED")
    print(f"Method: {request.method}")
    print(f"FILES keys: {request.FILES.keys()}")
    print(f"POST keys: {request.POST.keys()}")
    print("="*50)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Method not allowed'})
    
    User = get_user_model()
    
    try:
        # Ambil data dari session
        step1_data = request.session.get('step1_data', {})
        step2_data = request.session.get('step2_data', {})
        
        if not step1_data or not step2_data:
            messages.error(request, 'Session expired. Please start over.')
            return redirect('register_step', step=1)
        
        # File
        face_photos = request.FILES.getlist('foto_wajah')
        id_card_photo = request.FILES.get('foto_kartu_identitas')
        
        if len(face_photos) < 10:
            messages.error(request, f'Minimal 10 foto wajah diperlukan. Anda hanya upload {len(face_photos)}.')
            return redirect('register_step', step=3)
        
        if not id_card_photo:
            messages.error(request, 'Foto kartu identitas wajib diupload.')
            return redirect('register_step', step=3)
        
        # Validasi file
        valid_face_files = []
        for file in face_photos:
            if file.size > 5 * 1024 * 1024:
                messages.error(request, f'File {file.name} terlalu besar')
                return redirect('register_step', step=3)
            if not file.name.lower().endswith(('.jpg', '.jpeg', '.png')):
                messages.error(request, f'Format {file.name} tidak didukung')
                return redirect('register_step', step=3)
            valid_face_files.append(file)
        
        nim = step1_data.get('nim')
        email = step1_data.get('email')
        
        # ================= CEK EXISTING =================
        existing_user = User.objects.filter(username=nim).first()
        
        if existing_user:
            if existing_user.status_akun == 'pending':
                messages.error(request, f'NIM {nim} masih dalam proses verifikasi.')
                return redirect('register_step', step=1)
            elif existing_user.status_akun == 'aktif' and existing_user.is_active:
                messages.error(request, f'NIM {nim} sudah terdaftar dan aktif.')
                return redirect('register_step', step=1)
        
        existing_email = User.objects.filter(email=email).first()
        if existing_email and existing_email != existing_user:
            if existing_email.status_akun == 'pending':
                messages.error(request, f'Email {email} masih dalam proses verifikasi.')
                return redirect('register_step', step=1)
            elif existing_email.status_akun == 'aktif' and existing_email.is_active:
                messages.error(request, f'Email {email} sudah terdaftar dan aktif.')
                return redirect('register_step', step=1)
        
        # ================= TRANSACTION =================
        with transaction.atomic():

            # ===== HAPUS USER LAMA (DITOLAK) =====
            if existing_user and existing_user.status_akun == 'ditolak':
                print(f"DELETE user lama: {nim}")
                
                try:
                    old_mahasiswa = Mahasiswa.objects.get(user=existing_user)

                    old_mahasiswa.foto_wajah.all().delete()

                    if hasattr(old_mahasiswa, 'pengajuan_pendaftaran'):
                        old_mahasiswa.pengajuan_pendaftaran.delete()

                    old_mahasiswa.delete()

                except Mahasiswa.DoesNotExist:
                    pass

                existing_user.delete()
                existing_user = None  # 🔥 penting

            # ===== HAPUS EMAIL LAMA JIKA DITOLAK =====
            if existing_email and existing_email.status_akun == 'ditolak':
                print(f"DELETE email lama: {email}")
                
                try:
                    old_mahasiswa = Mahasiswa.objects.get(user=existing_email)

                    old_mahasiswa.foto_wajah.all().delete()

                    if hasattr(old_mahasiswa, 'pengajuan_pendaftaran'):
                        old_mahasiswa.pengajuan_pendaftaran.delete()

                    old_mahasiswa.delete()

                except Mahasiswa.DoesNotExist:
                    pass

                existing_email.delete()

            # ===== SAFETY NET (ANTI NYANGKUT) =====
            User.objects.filter(username=nim).delete()

            # ===== CREATE USER BARU =====
            user = User.objects.create_user(
                username=nim,
                email=email,
                password=step1_data.get('password'),
                nama_lengkap=step1_data.get('nama_lengkap'),
                nrp=nim,
                role='mahasiswa',
                status_akun='pending',
                is_active=False
            )

            # ===== FK =====
            jenjang = Jenjang_Pendidikan.objects.filter(id=step2_data.get('jenjang_id')).first()
            prodi_obj = Prodi.objects.filter(id=step2_data.get('prodi_id')).first()
            semester = Semester.objects.filter(id=step2_data.get('semester')).first()
            kelas_obj = Kelas.objects.filter(id=step2_data.get('kelas_id')).first()

            # ===== MAHASISWA =====
            mhs = Mahasiswa.objects.create(
                user=user,
                nim=nim,
                jenjang_pendidikan=jenjang,
                prodi=prodi_obj,
                semester=semester,
                kelas=kelas_obj,
                jurusan=step2_data.get('prodi_nama', ''),
                sks_total_tempuh=0,
                foto_ktm=id_card_photo
            )

            # ===== DOSEN =====
            dosen_ids = [
                (step2_data.get('dosen_pembimbing1'), 'pembimbing1'),
                (step2_data.get('dosen_pembimbing2'), 'pembimbing2'),
                (step2_data.get('dosen_pembimbing3'), 'pembimbing3'),
            ]

            for dosen_id, tipe in dosen_ids:
                if dosen_id:
                    d_obj = Dosen.objects.filter(id=dosen_id).first()
                    if d_obj:
                        Mahasiswa_Dosen.objects.create(
                            mahasiswa=mhs,
                            dosen=d_obj,
                            tipe_pembimbing=tipe
                        )

            # ===== KEGIATAN PA =====
            kp_ids = step2_data.get('kegiatan_pa_diambil', [])
            if kp_ids:
                kegiatan_objects = Kegiatan_PA.objects.filter(id__in=kp_ids)
                mhs.kegiatan_pa.set(kegiatan_objects)

                for kp in kegiatan_objects:
                    Status_Pemenuhan_SKS.objects.create(
                        mahasiswa=mhs,
                        kegiatan_pa=kp,
                        jam_target=kp.target_jam,
                        jumlah_sks=kp.jumlah_sks
                    )

            # ===== FOTO WAJAH =====
            for i, file_gambar in enumerate(valid_face_files):
                FotoWajah.objects.create(
                    mahasiswa=mhs,
                    file_path=file_gambar,
                    keterangan=f"Foto registrasi ke-{i+1}"
                )

            # ===== PENGAJUAN =====
            Pengajuan_Pendaftaran.objects.create(
                mahasiswa=mhs,
                status_pengajuan='pending'
            )

            # ===== SESSION =====
            request.session['registrasi_email'] = email
            request.session['registrasi_nama'] = step1_data.get('nama_lengkap')

            request.session.pop('step1_data', None)
            request.session.pop('step2_data', None)

            messages.success(request, 'Pendaftaran berhasil! Menunggu approval admin.')
            return redirect('registrasi_complete')

    except Exception as e:
        print(f"ERROR submit_registration: {e}")
        import traceback
        traceback.print_exc()
        messages.error(request, f'Terjadi kesalahan: {str(e)}')
        return redirect('register_step', step=3)
    
def get_registration_summary(request):
    """API untuk mendapatkan ringkasan data registrasi"""
    try:
        step1_data = request.session.get('step1_data', {})
        step2_data = request.session.get('step2_data', {})
        
        return JsonResponse({
            'success': True,
            'step1': {
                'nama_lengkap': step1_data.get('nama_lengkap', ''),
                'nim': step1_data.get('nim', ''),
                'email': step1_data.get('email', ''),
                'kelas_nama': step1_data.get('kelas_nama', ''),
            },
            'step2': {
                'prodi_nama': step2_data.get('prodi_nama', ''),
                'jenjang_nama': step2_data.get('jenjang_nama', ''),
                'semester_nama': step2_data.get('semester_nama', ''),
                'kegiatan_pa_list': ', '.join(step2_data.get('kegiatan_pa_nama', [])),
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

def login_view(request):
    # HAPUS CAPTCHA LAMA DARI SESSION (hanya untuk GET request pertama)
    if request.method == 'GET':
        if request.session.session_key:
            cache_key = f'captcha_{request.session.session_key}'
            cache.delete(cache_key)
    
    # Generate CAPTCHA untuk session
    if not request.session.session_key:
        request.session.create()
    
    if request.method == 'POST':
        username = request.POST.get("username")
        password = request.POST.get("password")
        captcha_input = request.POST.get("captcha_input", "")
        
        # VALIDASI CAPTCHA (TIDAK menghapus cache)
        is_captcha_valid = verify_captcha(request.session.session_key, captcha_input)
        
        if not is_captcha_valid:
            messages.error(request, 'CAPTCHA salah. Silakan coba lagi.')
            # JANGAN hapus CAPTCHA, biarkan user coba lagi dengan CAPTCHA yang sama
            return redirect('login')
        
        # Autentikasi user
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            # Cek status akun pending
            if getattr(user, 'status_akun', '') == 'pending':
                messages.warning(request, 'Akun Anda masih menunggu verifikasi admin.')
                # CAPTCHA valid tapi akun pending - HAPUS CAPTCHA (sudah terpakai)
                cache_key = f'captcha_{request.session.session_key}'
                cache.delete(cache_key)
                return redirect('login')
            
            # LOGIN BERHASIL - HAPUS CAPTCHA
            cache_key = f'captcha_{request.session.session_key}'
            cache.delete(cache_key)
            
            # Login user
            login(request, user)
            
            # Redirect berdasarkan role
            if getattr(user, 'role', '') == 'mahasiswa':
                return redirect('profil_mahasiswa')
            elif getattr(user, 'role', '') == 'admin' or user.is_superuser:
                return redirect('admin_dashboard')
            else:
                return redirect('login')
        else:
            # AUTHENTICATION GAGAL - JANGAN HAPUS CAPTCHA
            messages.error(request, 'Login Gagal. Cek kembali username dan password.')
            # Biarkan CAPTCHA tetap ada untuk percobaan berikutnya
            return redirect('login')
    
    # GET request - buat CAPTCHA baru (force refresh)
    captcha_data = get_or_create_captcha(request.session.session_key, force_refresh=True)
    
    return render(request, 'login.html', {
        'captcha_image': captcha_data['image']
    })
    
def registrasi_complete(request):
    """Halaman konfirmasi setelah registrasi selesai"""
    # Ambil email dari session
    email = request.session.get('registrasi_email', 'email.anda@domain.com')

    # Optional: Hapus session data setelah ditampilkan
    # request.session.pop('registrasi_email', None)

    return render(request, 'mahasiswa/persetujuandaftar.html', {'user_email': email})

@login_required
def edit_profil(request, nim):
    if request.method == 'POST':
        try:
            # Validasi bahwa NIM sesuai dengan user yang login
            if str(request.user.nrp) != str(nim) and request.user.role != 'admin':
                return JsonResponse({'success': False, 'error': 'Unauthorized'})
            
            user = request.user
            mahasiswa = Mahasiswa.objects.get(user=user)
            
            # Simpan kegiatan PA lama sebelum update
            old_kegiatan_ids = set(mahasiswa.kegiatan_pa.values_list('id', flat=True))
            
            # Update user data
            user.nama_lengkap = request.POST.get('nama')
            user.email = request.POST.get('email')
            
            # Update password jika diisi
            password = request.POST.get('password')
            confirm_password = request.POST.get('confirm_password')
            
            if password:
                if len(password) < 6:
                    return JsonResponse({'success': False, 'error': 'Password minimal 6 karakter'})
                if password != confirm_password:
                    return JsonResponse({'success': False, 'error': 'Password tidak sama'})
                user.set_password(password)
            
            user.save()
            
            # Update kelas
            kelas_id = request.POST.get('kelas')
            if kelas_id:
                try:
                    kelas_obj = Kelas.objects.get(id=kelas_id)
                    mahasiswa.kelas = kelas_obj
                except Kelas.DoesNotExist:
                    return JsonResponse({'success': False, 'error': 'Kelas tidak valid'})
            
            # Update semester
            semester_id = request.POST.get('semester')
            if semester_id:
                try:
                    semester_obj = Semester.objects.get(id=semester_id)
                    mahasiswa.semester = semester_obj
                except Semester.DoesNotExist:
                    return JsonResponse({'success': False, 'error': 'Semester tidak valid'})
            
            # Update jenjang
            jenjang_id = request.POST.get('jenjang')
            if jenjang_id:
                try:
                    jenjang = Jenjang_Pendidikan.objects.get(id=jenjang_id)
                    mahasiswa.jenjang_pendidikan = jenjang
                except Jenjang_Pendidikan.DoesNotExist:
                    return JsonResponse({'success': False, 'error': 'Jenjang tidak valid'})
            
            mahasiswa.save()
            
            # Update kegiatan PA (jika ada)
            kegiatan_pa_selected = request.POST.get('kegiatan_pa_selected')
            if kegiatan_pa_selected:
                try:
                    kegiatan_ids = json.loads(kegiatan_pa_selected)
                    new_kegiatan_ids = set(kegiatan_ids)
                    kegiatan_objects = Kegiatan_PA.objects.filter(id__in=kegiatan_ids)
                    
                    # Validasi minimal 1 kegiatan
                    if len(kegiatan_ids) == 0:
                        return JsonResponse({'success': False, 'error': 'Minimal pilih 1 kegiatan PA'})
                    
                    # CEK APAKAH KEGIATAN PA BERUBAH
                    kegiatan_changed = (old_kegiatan_ids != new_kegiatan_ids)
                    
                    if kegiatan_changed:
                        print(f"⚠️ Kegiatan PA berubah! Old: {old_kegiatan_ids}, New: {new_kegiatan_ids}")
                        print(f"🔄 Mereset progress SKS untuk mahasiswa {mahasiswa.user.nama_lengkap}")
                        
                        # HAPUS SEMUA PRESENSI LAMA
                        deleted_count, _ = Presensi.objects.filter(mahasiswa=mahasiswa).delete()
                        print(f"🗑️ Menghapus {deleted_count} presensi lama")
                        
                        # HAPUS SEMUA VERIFICATION LOGS
                        VerificationLog.objects.filter(mahasiswa=mahasiswa).delete()
                        
                        # HAPUS SEMUA DURASI
                        Durasi.objects.filter(presensi__mahasiswa=mahasiswa).delete()
                        
                        # RESET Status_Pemenuhan_SKS (akan dibuat ulang dengan kegiatan baru)
                        Status_Pemenuhan_SKS.objects.filter(mahasiswa=mahasiswa).delete()
                        
                        # Update kegiatan PA yang dipilih
                        mahasiswa.kegiatan_pa.set(kegiatan_objects)
                        
                        # Buat Status_Pemenuhan_SKS baru untuk setiap kegiatan
                        for kegiatan in kegiatan_objects:
                            Status_Pemenuhan_SKS.objects.create(
                                mahasiswa=mahasiswa,
                                kegiatan_pa=kegiatan,
                                jam_target=kegiatan.target_jam,
                                jumlah_sks=kegiatan.jumlah_sks,
                                jam_tercapai=0,  # Reset ke 0
                                status_pemenuhan='belum memenuhi'
                            )
                        
                        print(f"✅ Progress SKS direset, status baru dibuat untuk {len(kegiatan_objects)} kegiatan")
                        
                    else:
                        # Tidak ada perubahan kegiatan, update normal
                        mahasiswa.kegiatan_pa.set(kegiatan_objects)
                        
                        # Update Status_Pemenuhan_SKS untuk kegiatan yang masih ada
                        for kegiatan in kegiatan_objects:
                            Status_Pemenuhan_SKS.objects.update_or_create(
                                mahasiswa=mahasiswa,
                                kegiatan_pa=kegiatan,
                                defaults={
                                    'jam_target': kegiatan.target_jam,
                                    'jumlah_sks': kegiatan.jumlah_sks
                                }
                            )
                        
                        # Hapus status untuk kegiatan yang tidak dipilih lagi
                        Status_Pemenuhan_SKS.objects.filter(
                            mahasiswa=mahasiswa
                        ).exclude(
                            kegiatan_pa__in=kegiatan_objects
                        ).delete()
                    
                except json.JSONDecodeError as e:
                    print(f"JSON decode error: {e}")
                    return JsonResponse({'success': False, 'error': 'Data kegiatan tidak valid'})
            
            # Re-autentikasi user jika password diubah
            if password:
                from django.contrib.auth import update_session_auth_hash
                update_session_auth_hash(request, user)
            
            message = 'Profil berhasil diperbarui'
            if 'kegiatan_changed' in locals() and kegiatan_changed:
                message = 'Profil berhasil diperbarui. Progress SKS telah direset ke 0 karena perubahan kegiatan PA.'
            
            return JsonResponse({'success': True, 'message': message})
            
        except Mahasiswa.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Data mahasiswa tidak ditemukan'})
        except Exception as e:
            import traceback
            print(f"Error saving profile: {e}")
            traceback.print_exc()
            return JsonResponse({'success': False, 'error': f'Terjadi kesalahan: {str(e)}'})
    
    # Jika GET request, tetap render template
    return render(request, 'mahasiswa/edit_profil.html')

@login_required
def data_wajah(request):
    try:
        mahasiswa = Mahasiswa.objects.get(user=request.user)
        foto_wajah_list = FotoWajah.objects.filter(mahasiswa=mahasiswa)
        foto_count = foto_wajah_list.count()
        
        context = {
            'foto_wajah_list': foto_wajah_list,
            'foto_count': foto_count,
        }
        return render(request, 'mahasiswa/data_wajah.html', context)
    except Mahasiswa.DoesNotExist:
        messages.error(request, 'Data mahasiswa tidak ditemukan')
        return redirect('profil_mahasiswa')
    
@login_required
def hapus_foto_wajah(request, foto_id):
    if request.method == 'DELETE':
        try:
            user = request.user
            mahasiswa = Mahasiswa.objects.get(user=user)
            foto = FotoWajah.objects.get(id=foto_id, mahasiswa=mahasiswa)
            foto.delete()
            
            return JsonResponse({
                'success': True,
                'message': 'Foto berhasil dihapus'
            })
        except FotoWajah.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Foto tidak ditemukan'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

@login_required
def hapus_semua_foto(request):
    if request.method == 'POST':
        try:
            user = request.user
            mahasiswa = Mahasiswa.objects.get(user=user)
            foto_count = FotoWajah.objects.filter(mahasiswa=mahasiswa).count()
            FotoWajah.objects.filter(mahasiswa=mahasiswa).delete()
            
            return JsonResponse({
                'success': True,
                'message': f'{foto_count} foto berhasil dihapus'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

@login_required
def riwayat_presensi(request):
    try:
        mahasiswa = Mahasiswa.objects.get(user=request.user)
        
        # Ambil semua presensi mahasiswa ini
        presensi_qs = Presensi.objects.filter(
            mahasiswa=mahasiswa
        ).order_by('-tanggal_presensi', '-jam_checkin')
        
        presensi_list = []
        
        for p in presensi_qs:
            # Hitung durasi jika ada checkin dan checkout
            durasi_text = "-"
            if p.jam_checkin and p.jam_checkout:
                try:
                    # Gabungkan tanggal dengan waktu
                    checkin_datetime = datetime.combine(p.tanggal_presensi, p.jam_checkin)
                    checkout_datetime = datetime.combine(p.tanggal_presensi, p.jam_checkout)
                    
                    # Jika checkout lebih kecil dari checkin (lewat tengah malam)
                    if checkout_datetime < checkin_datetime:
                        checkout_datetime += timedelta(days=1)
                    
                    # Hitung selisih
                    delta = checkout_datetime - checkin_datetime
                    total_seconds = delta.total_seconds()
                    hours = int(total_seconds // 3600)
                    minutes = int((total_seconds % 3600) // 60)
                    
                    # Format durasi
                    if hours > 0:
                        durasi_text = f"{hours}j {minutes}m"
                    elif minutes > 0:
                        durasi_text = f"{minutes}m"
                    else:
                        durasi_text = "0m"
                        
                except Exception as e:
                    durasi_text = "Error"
                    print(f"ERROR hitung durasi: {e}")
            
            # Debug: Cetak waktu checkout
            print(f"DEBUG - Presensi {p.id}: Checkout={p.jam_checkout}, Session Status={p.session_status}")
            
            # Tambahkan ke list
            presensi_list.append({
                'tanggal': p.tanggal_presensi,
                'check_in': p.jam_checkin,
                'check_out': p.jam_checkout,
                'durasi': durasi_text,
                'session_status': p.session_status,
                'foto_checkin': p.foto_checkin.url if p.foto_checkin else None,
                'foto_checkout': p.foto_checkout.url if p.foto_checkout else None,
            })
        
        context = {
            'presensi_list': presensi_list,
            'mahasiswa': mahasiswa,
        }
        
        return render(request, 'mahasiswa/riwayat_presensi.html', context)
        
    except Mahasiswa.DoesNotExist:
        messages.error(request, 'Data mahasiswa tidak ditemukan')
        return redirect('profil_mahasiswa')
    except Exception as e:
        print(f"ERROR in riwayat_presensi: {str(e)}")
        import traceback
        traceback.print_exc()
        messages.error(request, 'Terjadi kesalahan saat mengambil data presensi')
        return redirect('profil_mahasiswa')

def logout_view(request):
    logout(request) 
    return redirect('login') 

# --- DASHBOARD VIEWS ---
@login_required
def kamera_presensi_mhs(request):
    """View untuk kamera presensi mahasiswa - hanya tampilkan yang sudah di-approve"""
    today = date.today()
    
    # Filter hanya mahasiswa yang status pengajuannya DISETUJUI
    mahasiswa_list = Mahasiswa.objects.select_related(
        'user',
        'jenjang_pendidikan',
        'semester'
    ).filter(  # hanya yang aktif
        pengajuan_pendaftaran__status_pengajuan='disetujui'  # hanya yang sudah di-approve
    ).order_by('user__nama_lengkap')
    
    # Debug: Cetak jumlah mahasiswa ke console
    print(f"DEBUG [Kamera Presensi]: Jumlah mahasiswa (approved): {mahasiswa_list.count()}")
    
    presensi_data = []
    for mahasiswa in mahasiswa_list:
        # Cari presensi hari ini
        presensi_today = Presensi.objects.filter(
            mahasiswa=mahasiswa,
            tanggal_presensi=today
        ).order_by('-jam_checkin')
        
        latest_presensi = None
        is_checked_in = False
        
        if presensi_today.exists():
            latest_presensi = presensi_today.first()
            # Cek apakah ada yang belum check-out
            pending_checkout = presensi_today.filter(
                jam_checkin__isnull=False,
                jam_checkout__isnull=True
            ).exists()
            is_checked_in = pending_checkout
        
        presensi_data.append({
            'mahasiswa': mahasiswa,
            'presensi': latest_presensi,
            'is_checked_in': is_checked_in,
            'all_presensi_today': presensi_today
        })
    
    context = {
        'presensi_data': presensi_data,
        'today': today,
        'total_mahasiswa': mahasiswa_list.count()
    }
    
    return render(request, 'admin/kamera_presensi_mhs.html', context)

@csrf_exempt
@login_required
def checkin_presensi(request):
    """
    API untuk check-in mahasiswa dengan liveness detection + face recognition
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            mahasiswa_id = data.get('mahasiswa_id')
            foto_base64 = data.get('foto')
            
            print(f"\n🔴 [checkin_presensi] DIPANGGIL!")
            print(f"   Mahasiswa ID: {mahasiswa_id}")
            print(f"   Timestamp: {datetime.now()}")
            
            if not mahasiswa_id or not foto_base64:
                return JsonResponse({
                    'success': False,
                    'message': 'Data tidak lengkap'
                })
            
            # Decode base64 image
            try:
                format, imgstr = foto_base64.split(';base64,')
                ext = format.split('/')[-1]
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f'checkin_{mahasiswa_id}_{timestamp}.{ext}'
                foto_data = ContentFile(base64.b64decode(imgstr), name=filename)
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'message': f'Gagal decode foto: {str(e)}'
                })
            
            mahasiswa = get_object_or_404(Mahasiswa, id=mahasiswa_id)
            today = date.today()
            
            # CEK SESSION AKTIF
            existing_presensi = Presensi.objects.filter(
                mahasiswa=mahasiswa,
                tanggal_presensi=today,
                jam_checkout__isnull=True
            ).exclude(session_status='completed')
            
            if existing_presensi.exists():
                if existing_presensi.count() > 1:
                    print(f"⚠️ Terdeteksi {existing_presensi.count()} session aktif! Membersihkan...")
                    latest = existing_presensi.order_by('-jam_checkin').first()
                    now_local = datetime.now()
                    for session in existing_presensi.exclude(id=latest.id):
                        session.jam_checkout = now_local.time()
                        session.session_status = 'auto_checkout'
                        session.save()
                    
                    active_session = latest
                else:
                    active_session = existing_presensi.first()
                
                return JsonResponse({
                    'success': False,
                    'message': 'Anda masih memiliki sesi aktif. Silakan checkout terlebih dahulu.',
                    'existing_session': {
                        'id': active_session.id,
                        'jam_checkin': active_session.jam_checkin.strftime('%H:%M'),
                        'session_status': active_session.session_status
                    }
                })
            
            # SIMPAN FOTO
            foto_wajah = FotoWajah.objects.create(
                mahasiswa=mahasiswa,
                file_path=foto_data,
                keterangan=f'Check-in {datetime.now().strftime("%d/%m/%Y %H:%M")}'
            )
            
            now_local = datetime.now()
            
            # HITUNG DEADLINE (10 MENIT DARI SEKARANG)
            monitoring_deadline = now_local + timedelta(minutes=10)
            
            # BUAT PRESENSI dengan status 'waiting_monitoring'
            presensi = Presensi.objects.create(
                mahasiswa=mahasiswa,
                kegiatan_pa=None,
                tanggal_presensi=today,
                jam_checkin=now_local.time(),
                foto_checkin=foto_data,
                last_verified_at=now_local,
                terakhir_terdeteksi=now_local.time(),
                consecutive_failures=0,
                session_status='waiting_monitoring',  # PENTING!
                monitoring_deadline=monitoring_deadline,  # PENTING!
                sudah_verifikasi=False
            )
            
            VerificationLog.objects.create(
                mahasiswa=mahasiswa,
                presensi=presensi,
                timestamp=datetime.now(),
                status=True,
                is_liveness_real=True,
                failure_count=0,
                foto=foto_data
            )
            
            Durasi.objects.create(
                presensi=presensi,
                waktu_durasi=timedelta(seconds=0)
            )
            
            print(f"\n=== CHECK-IN BERHASIL ===")
            print(f"Mahasiswa: {mahasiswa.user.nama_lengkap} ({mahasiswa.nim})")
            print(f"Presensi ID: {presensi.id}")
            print(f"Session Status: waiting_monitoring")
            print(f"Monitoring Deadline: {monitoring_deadline.strftime('%H:%M:%S')}")
            print(f"==========================\n")
            
            return JsonResponse({
                'success': True,
                'message': 'Check-in berhasil! Anda memiliki waktu 10 menit untuk memulai monitoring.',
                'data': {
                    'presensi_id': presensi.id,
                    'jam_checkin': presensi.jam_checkin.strftime('%H:%M'),
                    'session_status': presensi.session_status,
                    'monitoring_deadline': monitoring_deadline.isoformat(),
                    'consecutive_failures': presensi.consecutive_failures,
                    'last_verified_at': presensi.last_verified_at.isoformat() if presensi.last_verified_at else None
                }
            })
            
        except Mahasiswa.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Mahasiswa tidak ditemukan'
            })
        except Exception as e:
            import traceback
            print(f"ERROR checkin_presensi: {str(e)}")
            traceback.print_exc()
            return JsonResponse({
                'success': False,
                'message': f'Terjadi kesalahan: {str(e)}'
            })
    
    return JsonResponse({
        'success': False,
        'message': 'Method tidak diizinkan'
    })

@login_required
def debug_presensi_data(request):
    """Fungsi untuk debugging data presensi"""
    try:
        mahasiswa = Mahasiswa.objects.get(user=request.user)
        
        # Ambil semua presensi
        presensi_list = Presensi.objects.filter(
            mahasiswa=mahasiswa,
            jam_checkin__isnull=False,
            jam_checkout__isnull=False
        ).order_by('-tanggal_presensi')
        
        data = []
        total_hours = 0
        
        for presensi in presensi_list:
            checkin_dt = datetime.combine(presensi.tanggal_presensi, presensi.jam_checkin)
            checkout_dt = datetime.combine(presensi.tanggal_presensi, presensi.jam_checkout)
            
            if checkout_dt < checkin_dt:
                checkout_dt += timedelta(days=1)
            
            durasi = checkout_dt - checkin_dt
            hours = durasi.total_seconds() / 3600
            total_hours += hours
            
            data.append({
                'tanggal': presensi.tanggal_presensi,
                'checkin': presensi.jam_checkin,
                'checkout': presensi.jam_checkout,
                'durasi': str(durasi),
                'hours': round(hours, 2)
            })
        
        return JsonResponse({
            'success': True,
            'total_presensi': len(data),
            'total_hours': round(total_hours, 2),
            'data': data,
            'mahasiswa': {
                'nama': mahasiswa.user.nama_lengkap,
                'nim': mahasiswa.nim
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@csrf_exempt
def detect_liveness_frame(request):
    """
    API endpoint untuk deteksi liveness - UPDATED VERSION
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'})
    
    try:
        data = json.loads(request.body)
        frame_base64 = data.get('frame')
        mahasiswa_id = data.get('mahasiswa_id')
        action = data.get('action')
        
        if not frame_base64:
            return JsonResponse({
                'success': False, 
                'error': 'No frame provided', 
                'status': 'ERROR'
            })
        
        # Gunakan mahasiswa_id sebagai session_id untuk state terpisah per user
        session_id = str(mahasiswa_id) if mahasiswa_id else "default"
        
        # Panggil liveness detection dengan session_id
        from liveness_detection import process_frame_liveness, reset_detection_state
        result = process_frame_liveness(frame_base64, session_id=session_id)
        
        # ====================== HANDLE SPOOF ======================
        if result.get('status') == 'SPOOF':
            print(f"[SPOOF] ❌ Spoof detected for session {session_id}")
            return JsonResponse({
                **result,
                'stop_detection': True,
                'force_close_modal': True,
                'should_save': False,
                'message': 'SPOOF DETECTED! Please use your real face.'
            })
        
        # ====================== HANDLE NO FACE ======================
        if result.get('status') == 'NO_FACE':
            return JsonResponse({
                **result,
                'verified': False,
                'should_save': False
            })
        
        # ====================== HANDLE LIVENESS_CHECK ======================
        if result.get('status') == 'LIVENESS_CHECK':
            return JsonResponse({
                **result,
                'verified': False,
                'should_save': False
            })
        
        # ====================== HANDLE REAL ======================
        if result.get('status') == 'REAL' and result.get('verified'):
            # Validasi mahasiswa_id
            if not mahasiswa_id:
                return JsonResponse({
                    'success': True,
                    'status': 'ERROR',
                    'message': 'No student selected',
                    'verified': False,
                    'should_save': False
                })
            
            # Lakukan face recognition
            from .face_recognition_utils import verify_face_with_insightface
            
            box = result.get('box')
            recognition_result = verify_face_with_insightface(
                frame_base64, 
                mahasiswa_id, 
                face_box=box
            )
            
            # Face tidak cocok
            if not recognition_result.get('verified'):
                print(f"[RECOGNITION] ❌ Face mismatch for mahasiswa {mahasiswa_id}")
                reset_detection_state(session_id)
                return JsonResponse({
                    'success': True,
                    'status': 'NOT_RECOGNIZED',
                    'message': 'Wajah tidak cocok dengan data mahasiswa!',
                    'verified': False,
                    'blink_count': result.get('blink_count', 0),
                    'spoof_score': result.get('spoof_score', 0),
                    'face_detected': True,
                    'box': result.get('box'),
                    'stop_detection': True,
                    'force_close_modal': True,
                    'should_save': False
                })
            
            # Face cocok - lanjut ke save
            mahasiswa = get_object_or_404(Mahasiswa, id=mahasiswa_id)
            today = date.today()
            
            result['recognition_score'] = recognition_result.get('recognition_score', 0)
            result['nama_mahasiswa'] = mahasiswa.user.nama_lengkap or mahasiswa.user.username
            result['verified'] = True
            result['should_save'] = True
            result['mahasiswa_id'] = mahasiswa_id
            
            print(f"[RECOGNITION] ✅ Face matched for {result['nama_mahasiswa']}")
            
            # ====================== 🔥 CEK SESSION SEBELUM SAVE 🔥 ======================
            existing_presensi = Presensi.objects.filter(
                mahasiswa=mahasiswa,
                tanggal_presensi=today,
                jam_checkout__isnull=True
            ).exclude(session_status='completed')
            
            # ====================== HANDLE CHECK-IN ======================
            if action == 'checkin':
                # 🔥 CEK APAKAH SUDAH ADA SESSION AKTIF 🔥
                if existing_presensi.exists():
                    active_session = existing_presensi.first()
                    print(f"[SESSION] Active session exists for {result['nama_mahasiswa']}, cannot check-in again")
                    return JsonResponse({
                        'success': True,
                        'status': 'SESSION_ACTIVE',
                        'message': 'Anda masih memiliki sesi aktif! Silakan checkout terlebih dahulu.',
                        'verified': False,
                        'stop_detection': True,
                        'force_close_modal': True,
                        'should_save': False,
                        'existing_session': {
                            'id': active_session.id,
                            'jam_checkin': active_session.jam_checkin.strftime('%H:%M:%S'),
                            'session_status': active_session.session_status
                        }
                    })
                
                # 🔥 BUAT PRESENSI BARU (HANYA 1 KALI) 🔥
                foto_data = _save_base64_to_file(frame_base64, f'checkin_{mahasiswa_id}')
                now_local = datetime.now()
                
                presensi = Presensi.objects.create(
                    mahasiswa=mahasiswa,
                    kegiatan_pa=None,
                    tanggal_presensi=today,
                    jam_checkin=now_local.time(),
                    foto_checkin=foto_data,
                    last_verified_at=now_local,
                    terakhir_terdeteksi=now_local.time(),
                    consecutive_failures=0,
                    session_status='waiting_monitoring',
                    monitoring_deadline=now_local + timedelta(minutes=10),
                    sudah_verifikasi=False
                )
                
                VerificationLog.objects.create(
                    mahasiswa=mahasiswa,
                    presensi=presensi,
                    timestamp=datetime.now(),
                    status=True,
                    is_liveness_real=True,
                    failure_count=0,
                    foto=foto_data
                )
                
                Durasi.objects.create(
                    presensi=presensi,
                    waktu_durasi=timedelta(seconds=0)
                )
                
                result['presensi_id'] = presensi.id
                result['saved'] = True
                result['action'] = 'checkin'
                print(f"[SAVE] ✅ Check-in successful for {mahasiswa.user.nama_lengkap}")
            
            # ====================== HANDLE CHECK-OUT ======================
            elif action == 'checkout':
                # 🔥 CEK APAKAH ADA SESSION AKTIF 🔥
                if not existing_presensi.exists():
                    print(f"[SESSION] No active session for checkout for {result['nama_mahasiswa']}")
                    return JsonResponse({
                        'success': True,
                        'status': 'NO_ACTIVE_SESSION',
                        'message': 'Tidak ada sesi aktif untuk checkout! Silakan check-in terlebih dahulu.',
                        'verified': False,
                        'stop_detection': True,
                        'force_close_modal': True,
                        'should_save': False
                    })
                
                presensi = existing_presensi.first()
                
                # CEK JANGAN DOUBLE CHECKOUT
                if presensi.jam_checkout is not None:
                    print(f"[CHECKOUT] Session already checked out at {presensi.jam_checkout}")
                    return JsonResponse({
                        'success': True,
                        'status': 'ALREADY_CHECKED_OUT',
                        'message': 'Sesi sudah checkout sebelumnya!',
                        'verified': False,
                        'stop_detection': True,
                        'force_close_modal': True,
                        'should_save': False
                    })
                
                foto_data = _save_base64_to_file(frame_base64, f'checkout_{mahasiswa_id}')
                now_local = datetime.now()
                
                presensi.jam_checkout = now_local.time()
                presensi.foto_checkout = foto_data
                presensi.session_status = 'completed'
                presensi.save()
                
                # Hitung durasi
                checkin_dt = datetime.combine(today, presensi.jam_checkin)
                checkout_dt = datetime.combine(today, presensi.jam_checkout)
                if checkout_dt < checkin_dt:
                    checkout_dt += timedelta(days=1)
                durasi = checkout_dt - checkin_dt
                
                Durasi.objects.update_or_create(
                    presensi=presensi,
                    defaults={'waktu_durasi': durasi}
                )
                
                # Simpan foto ke FotoWajah untuk dataset
                FotoWajah.objects.create(
                    mahasiswa=mahasiswa,
                    file_path=foto_data,
                    keterangan=f'Check-out {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}'
                )
                
                result['presensi_id'] = presensi.id
                result['saved'] = True
                result['action'] = 'checkout'
                print(f"[SAVE] ✅ Check-out successful for {mahasiswa.user.nama_lengkap}")
            
            # Reset state setelah save berhasil
            reset_detection_state(session_id)
            
            return JsonResponse(result)
        
        # Default return
        return JsonResponse(result)
        
    except json.JSONDecodeError as e:
        print(f"[ERROR] JSON decode error: {e}")
        return JsonResponse({
            'success': False, 
            'error': 'Invalid JSON',
            'status': 'ERROR'
        })
    except Exception as e:
        print(f"[ERROR] detect_liveness_frame: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': str(e),
            'status': 'ERROR'
        })

@csrf_exempt
def detect_face_registration(request):
    """
    API endpoint khusus untuk deteksi wajah saat registrasi (TANPA side-effect liveness blink)
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            frame_base64 = data.get('frame')
            
            if not frame_base64:
                return JsonResponse({
                    'success': False, 
                    'error': 'No frame provided',
                    'face_detected': False
                })

            # Decode frame
            from liveness_detection import decode_base64_image, apply_gamma_correction, detector, init_liveness_detection, GAMMA_VALUE
            import cv2
            import numpy as np

            # Pastikan detector siap
            if detector is None:
                init_liveness_detection()
                from liveness_detection import detector as initialized_detector
                d = initialized_detector
            else:
                d = detector

            frame = decode_base64_image(frame_base64)
            if frame is None:
                return JsonResponse({
                    'success': False, 
                    'error': 'Failed to decode image',
                    'face_detected': False
                })
            
            # Apply gamma correction (supaya konsisten dengan liveness)
            frame_gamma = apply_gamma_correction(frame, GAMMA_VALUE)
            rgb = cv2.cvtColor(frame_gamma, cv2.COLOR_BGR2RGB)
            
            # Detect faces
            from liveness_detection import detector as d_fix
            if d_fix is None:
                init_liveness_detection()
                from liveness_detection import detector as d_fix
            
            detections = d_fix.detect_faces(rgb)
            
            if len(detections) > 0:
                # Ambil face pertama dengan confidence cukup
                best_face = None
                for det in detections:
                    if det['confidence'] > 0.5:
                        best_face = det
                        break
                
                if best_face:
                    x, y, w, h = best_face['box']
                    return JsonResponse({
                        'success': True,
                        'face_detected': True,
                        'box': {'x': int(x), 'y': int(y), 'w': int(w), 'h': int(h)},
                        'frame_width': frame.shape[1],
                        'frame_height': frame.shape[0]
                    })

            return JsonResponse({
                'success': True,
                'face_detected': False,
                'message': 'No face detected'
            })

        except Exception as e:
            print(f"[ERROR] detect_face_registration: {e}")
            return JsonResponse({
                'success': False, 
                'error': str(e),
                'face_detected': False
            })

    return JsonResponse({'success': False, 'error': 'Method not allowed'})

@csrf_exempt
@login_required
def periodic_verify(request):
    """
    API endpoint untuk verifikasi periodik
    - Jika status 'waiting_monitoring': verifikasi pertama, jika berhasil pindah ke 'monitoring_active'
    - Jika lewat deadline 10 menit: AUTO CHECKOUT
    - Jika status 'monitoring_active': verifikasi periodik tiap 5 menit, auto checkout jika 2x gagal
    """
    if request.method != 'POST':
        return JsonResponse({
            "success": False, 
            "message": "Method not allowed"
        })
    
    try:
        data = json.loads(request.body)
        frame_base64 = data.get('frame')
        presensi_id = data.get('presensi_id')
        
        # Validasi user adalah mahasiswa
        if request.user.role != 'mahasiswa':
            return JsonResponse({
                "success": False, 
                "message": "Hanya mahasiswa yang bisa menggunakan fitur ini"
            })
        
        # Ambil data mahasiswa
        mahasiswa = get_object_or_404(Mahasiswa, user=request.user)
        today = date.today()
        
        # CARI SESSION AKTIF
        if presensi_id:
            presensi = get_object_or_404(
                Presensi, 
                id=presensi_id,
                mahasiswa=mahasiswa,
                tanggal_presensi=today,
                jam_checkout__isnull=True
            )
        else:
            presensi = Presensi.objects.filter(
                mahasiswa=mahasiswa,
                tanggal_presensi=today,
                jam_checkout__isnull=True
            ).exclude(
                session_status='completed'
            ).order_by('-jam_checkin').first()
        
        if not presensi:
            return JsonResponse({
                "success": False, 
                "message": "Tidak ada sesi presensi aktif. Silakan check-in terlebih dahulu.",
                "session_status": "none",
                "clear_storage": True
            })
        
        now = datetime.now()
        
        # ==================== CEK DEADLINE 10 MENIT ====================
        if presensi.session_status == 'waiting_monitoring':
            if presensi.monitoring_deadline and now > presensi.monitoring_deadline:
                print(f"[DEADLINE] ⚠️ Monitoring tidak dimulai dalam 10 menit! Auto checkout!")
                
                presensi.session_status = 'auto_checkout'
                presensi.jam_checkout = now.time()
                presensi.consecutive_failures = 0
                
                # Hitung durasi
                checkin_dt = datetime.combine(presensi.tanggal_presensi, presensi.jam_checkin)
                checkout_dt = datetime.combine(presensi.tanggal_presensi, presensi.jam_checkout)
                if checkout_dt < checkin_dt:
                    checkout_dt += timedelta(days=1)
                durasi = checkout_dt - checkin_dt
                
                Durasi.objects.update_or_create(
                    presensi=presensi,
                    defaults={'waktu_durasi': durasi}
                )
                
                presensi.save()
                
                return JsonResponse({
                    "success": False,
                    "message": "Auto checkout: Monitoring tidak dimulai dalam 10 menit",
                    "auto_checked_out": True,
                    "clear_storage": True,
                    "redirect": "/riwayat_presensi/"
                })
        
        # ==================== CEK APAKAH SUDAH TERLALU CEPAT ====================
        MIN_VERIFICATION_INTERVAL = 30
        if presensi.last_verified_at:
            time_since_last = (now - presensi.last_verified_at).total_seconds()
            if time_since_last < MIN_VERIFICATION_INTERVAL and time_since_last > 0:
                print(f"[VERIFY] ⏱️ Too soon! Only {time_since_last:.1f}s since last verification.")
                return JsonResponse({
                    "success": True,
                    "verified": False,
                    "message": f"Tunggu {int(MIN_VERIFICATION_INTERVAL - time_since_last)} detik",
                    "wait_time": int(MIN_VERIFICATION_INTERVAL - time_since_last),
                    "consecutive_failures": presensi.consecutive_failures,
                    "last_verified_at": presensi.last_verified_at.isoformat(),
                    "skip_verification": True
                })
        
        # ==================== LIVENESS DETECTION ====================
        from liveness_detection import (
            decode_base64_image, 
            apply_gamma_correction, 
            detector, 
            model, 
            IMG_SIZE, 
            SPOOF_THRESHOLD, 
            GAMMA_VALUE, 
            init_liveness_detection
        )
        import cv2
        import numpy as np
        from .face_recognition_utils import verify_face_with_insightface
        
        if detector is None or model is None:
            init_liveness_detection()
            from liveness_detection import detector as d_init, model as m_init
            d, m = d_init, m_init
        else:
            d, m = detector, model
        
        frame = decode_base64_image(frame_base64)
        if frame is None:
            return JsonResponse({
                "success": False, 
                "message": "Gagal decode gambar"
            })
        
        frame_gamma = apply_gamma_correction(frame, GAMMA_VALUE)
        rgb = cv2.cvtColor(frame_gamma, cv2.COLOR_BGR2RGB)
        detections = d.detect_faces(rgb)
        
        is_real = False
        verified = False
        face_box = None
        face_detected = len(detections) > 0
        multiple_faces = len(detections) > 1
        recognition_score = 0
        nama_mahasiswa = None
        
        verification_success = False
        
        if face_detected and not multiple_faces:
            best_det = max(detections, key=lambda x: x['confidence'])
            
            if best_det and best_det['confidence'] > 0.5:
                x, y, w, h = best_det['box']
                face_box = {'x': int(x), 'y': int(y), 'w': int(w), 'h': int(h)}
                face = frame_gamma[max(0, y):y+h, max(0, x):x+w]
                
                if face.size > 0:
                    face_resized = cv2.resize(face, (IMG_SIZE, IMG_SIZE))
                    face_input = face_resized / 255.0
                    face_input = np.expand_dims(face_input, axis=0)
                    spoof_score = float(m.predict(face_input, verbose=0)[0][0])
                    is_real = spoof_score >= SPOOF_THRESHOLD
                    
                    if is_real:
                        try:
                            recognition_result = verify_face_with_insightface(
                                frame_base64, 
                                mahasiswa.id, 
                                face_box=face_box
                            )
                            verified = recognition_result.get('verified', False)
                            recognition_score = recognition_result.get('recognition_score', 0)
                            nama_mahasiswa = recognition_result.get('nama_mahasiswa')
                        except Exception as e:
                            print(f"[PERIODIC] Recognition error: {e}")
                            verified = False
                
                verification_success = face_detected and not multiple_faces and is_real and verified
        
        # ==================== UPDATE DATA ====================
        now_local = datetime.now()
        
        # SIMPAN FOTO KE DATABASE
        timestamp_str = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename_verif = f'verif_{mahasiswa.nim}_{timestamp_str}.jpg'
        
        if ';base64,' in frame_base64:
            imgstr_verif = frame_base64.split(';base64,')[1]
        else:
            imgstr_verif = frame_base64
            
        foto_data_verif = ContentFile(base64.b64decode(imgstr_verif), name=filename_verif)
        
        presensi.last_verified_at = now_local
        
        if face_detected:
            presensi.terakhir_terdeteksi = now_local.time()
        
        auto_checked_out = False
        checkout_time_str = None
        message = ""
        
        # ==================== LOGIKA BERDASARKAN SESSION STATUS ====================
        
        if presensi.session_status == 'waiting_monitoring':
            # INI VERIFIKASI PERTAMA KALI
            if verification_success:
                # ✅ BERHASIL: Pindah ke monitoring_active
                presensi.session_status = 'monitoring_active'
                presensi.consecutive_failures = 0
                presensi.sudah_verifikasi = True
                message = "Verifikasi pertama berhasil! Monitoring dimulai."
                print(f"[VERIFY] ✅ First verification SUCCESS! Status berubah ke monitoring_active")
            else:
                # ❌ GAGAL: Catat failure
                presensi.consecutive_failures += 1
                message = f"Verifikasi gagal: {presensi.consecutive_failures} dari 2 kegagalan"
                print(f"[VERIFY] ❌ First verification FAILED! failures={presensi.consecutive_failures}")
                
                # Jika sudah 2x gagal dalam grace period, auto checkout
                if presensi.consecutive_failures >= 2:
                    print(f"[VERIFY] 🚨 2 failures during grace period! Auto checkout!")
                    presensi.session_status = 'auto_checkout'
                    presensi.jam_checkout = now_local.time()
                    presensi.foto_checkout = foto_data_verif
                    auto_checked_out = True
                    checkout_time_str = now_local.strftime('%H:%M:%S')
                    message = "Auto checkout: 2x kegagalan verifikasi"
        
        elif presensi.session_status == 'monitoring_active':
            # MONITORING AKTIF: Verifikasi periodik tiap 5 menit
            if verification_success:
                presensi.consecutive_failures = 0
                message = "Verifikasi berhasil"
                print(f"[VERIFY] ✅ SUCCESS - Reset failures to 0")
            else:
                presensi.consecutive_failures += 1
                print(f"[VERIFY] ❌ FAILED - failures = {presensi.consecutive_failures}")
                
                if not face_detected:
                    message = "Wajah tidak terdeteksi"
                elif multiple_faces:
                    message = "Terdeteksi lebih dari satu wajah"
                elif not is_real:
                    message = "Spoofing terdeteksi"
                elif not verified:
                    message = "Identitas wajah tidak cocok"
                else:
                    message = "Verifikasi gagal"
                
                if presensi.consecutive_failures >= 2:
                    print(f"[VERIFY] 🚨 AUTO CHECKOUT TRIGGERED - 2 consecutive failures!")
                    presensi.session_status = 'auto_checkout'
                    presensi.jam_checkout = now_local.time()
                    presensi.foto_checkout = foto_data_verif
                    auto_checked_out = True
                    checkout_time_str = now_local.strftime('%H:%M:%S')
                    message = f"Auto checkout: 2x kegagalan verifikasi berturut-turut"
        
        # Simpan perubahan
        update_fields = ['last_verified_at', 'consecutive_failures', 'session_status']
        if face_detected:
            update_fields.append('terakhir_terdeteksi')
        if auto_checked_out:
            update_fields.extend(['jam_checkout', 'foto_checkout'])
        if presensi.session_status == 'monitoring_active' and verification_success:
            update_fields.append('sudah_verifikasi')
        
        presensi.save(update_fields=update_fields)
        
        # ==================== BUAT VERIFICATION LOG ====================
        VerificationLog.objects.create(
            mahasiswa=mahasiswa,
            presensi=presensi,
            timestamp=datetime.now(),
            status=verification_success,
            is_liveness_real=is_real,
            failure_count=presensi.consecutive_failures,
            foto=foto_data_verif
        )
        
        # Hitung kapan next verification
        VERIFICATION_INTERVAL = 5 * 60
        next_verification_time = presensi.last_verified_at + timedelta(seconds=VERIFICATION_INTERVAL)
        time_until_next = max(0, int((next_verification_time - datetime.now()).total_seconds()))
        
        # ==================== RESPONSE ====================
        response_data = {
            "success": True,
            "verified": verification_success,
            "is_real": is_real,
            "face_detected": face_detected,
            "multiple_faces": multiple_faces,
            "consecutive_failures": presensi.consecutive_failures,
            "auto_checked_out": auto_checked_out,
            "session_status": presensi.session_status,
            "message": message,
            "presensi_id": presensi.id,
            "checkout_time": checkout_time_str,
            "last_verified_at": presensi.last_verified_at.isoformat(),
            "next_verification_at": next_verification_time.isoformat(),
            "time_until_next": time_until_next,
            "details": {
                "liveness": "REAL" if is_real else "SPOOF" if face_detected else "NO_FACE",
                "recognition": "MATCHED" if verification_success else "MISMATCH" if face_detected and is_real else "SKIPPED",
                "faces": len(detections) if detections is not None else 0,
                "recognition_score": recognition_score if verification_success else 0,
            }
        }
        
        if auto_checked_out:
            response_data["warning"] = "Anda telah di-checkout otomatis"
            response_data["redirect"] = "/riwayat_presensi/"
            response_data["clear_storage"] = True
        
        return JsonResponse(response_data)

    except Exception as e:
        import traceback
        print(f"[periodic_verify Error] {str(e)}")
        traceback.print_exc()
        return JsonResponse({
            "success": False, 
            "message": f"Terjadi kesalahan: {str(e)}",
            "auto_checked_out": False,
            "clear_storage": False
        })
        
@require_http_methods(["GET"])
@csrf_exempt
def get_semester_by_jenjang(request, jenjang_id):
    """API untuk mendapatkan semester berdasarkan jenjang"""
    try:
        jenjang = Jenjang_Pendidikan.objects.get(id=jenjang_id)
        jenjang_nama = jenjang.nama_jenjang
        
        if "D3" in jenjang_nama:
            jenjang_kode = 'D3'
        elif "D4" in jenjang_nama:
            jenjang_kode = 'D4'
        elif "S2" in jenjang_nama:
            jenjang_kode = 'S2'
        else:
            jenjang_kode = None
        
        if jenjang_kode:
            semesters = Semester.objects.filter(jenjang=jenjang_kode).order_by('nomor_semester')
            semester_list = [{'id': s.id, 'nama_semester': s.nama_semester} for s in semesters]
        else:
            semester_list = []
        
        return JsonResponse({'success': True, 'semesters': semester_list})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e), 'semesters': []})

@login_required
def get_session_status(request):
    """
    API untuk mendapatkan status session monitoring terbaru
    Dipanggil secara periodik oleh frontend (setiap 30 detik)
    """
    try:
        if request.user.role != 'mahasiswa':
            return JsonResponse({
                'success': False,
                'error': 'Hanya untuk mahasiswa'
            })
        
        mahasiswa = get_object_or_404(Mahasiswa, user=request.user)
        today = date.today()
        
        # Cari session aktif hari ini (belum checkout)
        active_session = Presensi.objects.filter(
            mahasiswa=mahasiswa,
            tanggal_presensi=today,
            jam_checkout__isnull=True
        ).exclude(
            session_status='completed'
        ).order_by('-jam_checkin').first()
        
        if not active_session:
            # Bersihkan session lama
            old_active = Presensi.objects.filter(
                mahasiswa=mahasiswa,
                jam_checkout__isnull=True
            ).exclude(
                tanggal_presensi=today
            ).order_by('-tanggal_presensi', '-jam_checkin').first()
            
            if old_active:
                print(f"⚠️ Menemukan session lama yang masih aktif: {old_active.id}")
                now_local = datetime.now()
                old_active.jam_checkout = now_local.time()
                old_active.session_status = 'auto_checkout'
                old_active.save()
                print(f"✅ Auto checkout session lama")
            
            return JsonResponse({
                'success': True,
                'has_active_session': False,
                'message': 'Tidak ada session aktif',
                'clear_storage': True
            })
        
        # 🔥🔥🔥 TAMBAHKAN INI: CEK DEADLINE DI API INI JUGA 🔥🔥🔥
        from django.utils import timezone
        now = timezone.now()
        
        if active_session.session_status == 'waiting_monitoring':
            if active_session.monitoring_deadline and now > active_session.monitoring_deadline:
                print(f"[AUTO CHECKOUT via API] Deadline lewat untuk session {active_session.id}")
                
                active_session.session_status = 'auto_checkout'
                active_session.jam_checkout = now.time()
                active_session.save()
                
                return JsonResponse({
                    'success': True,
                    'has_active_session': False,
                    'clear_storage': True,
                    'auto_checked_out': True,
                    'message': 'Sesi di-checkout karena tidak memulai monitoring dalam 10 menit'
                })
        
        # Bersihkan session duplikat
        other_active = Presensi.objects.filter(
            mahasiswa=mahasiswa,
            tanggal_presensi=today,
            jam_checkout__isnull=True
        ).exclude(id=active_session.id)
        
        if other_active.exists():
            print(f"⚠️ Menemukan {other_active.count()} session aktif lain")
            now_local = datetime.now()
            for session in other_active:
                session.jam_checkout = now_local.time()
                session.session_status = 'auto_checkout'
                session.save()
        
        import pytz
        
        WIB = pytz.timezone('Asia/Jakarta')
        
        def to_wib(dt):
            if dt is None:
                return None
            if timezone.is_naive(dt):
                return WIB.localize(dt)
            return dt.astimezone(WIB)
        
        now_wib = to_wib(now)
        checkin_dt = datetime.combine(active_session.tanggal_presensi, active_session.jam_checkin)
        checkin_wib = to_wib(checkin_dt)
        
        # Hitung durasi session
        duration_seconds = int((now_wib - checkin_wib).total_seconds())
        if duration_seconds < 0:
            duration_seconds = 0
        
        # HITUNG SISA WAKTU BERDASARKAN SESSION STATUS
        VERIFICATION_INTERVAL = 5 * 60
        GRACE_PERIOD_MINUTES = 10
        
        deadline_seconds_left = 0
        next_verification_seconds_left = 0
        
        if active_session.session_status == 'waiting_monitoring':
            if active_session.monitoring_deadline:
                deadline_seconds_left = max(0, int((active_session.monitoring_deadline - now).total_seconds()))
        
        elif active_session.session_status == 'monitoring_active':
            if active_session.last_verified_at:
                last_verified_wib = to_wib(active_session.last_verified_at)
                next_verification_wib = last_verified_wib + timedelta(seconds=VERIFICATION_INTERVAL)
                next_verification_seconds_left = max(0, int((next_verification_wib - now_wib).total_seconds()))
            else:
                next_verification_seconds_left = VERIFICATION_INTERVAL
        
        # Cek apakah sudah pernah verifikasi
        sudah_verifikasi = VerificationLog.objects.filter(
            presensi=active_session
        ).exists()
        
        session_data = {
            'id': active_session.id,
            'checkin_time': active_session.jam_checkin.strftime('%H:%M:%S'),
            'checkin_datetime': checkin_wib.isoformat(),
            'duration_seconds': duration_seconds,
            'consecutive_failures': active_session.consecutive_failures,
            'session_status': active_session.session_status,
            'last_verified_at': to_wib(active_session.last_verified_at).isoformat() if active_session.last_verified_at else None,
            'terakhir_terdeteksi': active_session.terakhir_terdeteksi.strftime('%H:%M:%S') if active_session.terakhir_terdeteksi else None,
            'sudah_verifikasi': sudah_verifikasi,
            'deadline_seconds_left': deadline_seconds_left,
            'next_verification_seconds_left': next_verification_seconds_left,
            'monitoring_deadline': active_session.monitoring_deadline.isoformat() if active_session.monitoring_deadline else None,
        }
        
        # Ambil 10 log terakhir
        recent_logs = VerificationLog.objects.filter(
            presensi=active_session
        ).order_by('-timestamp')[:10]

        logs_data = []
        for log in recent_logs:
            log_time_wib = to_wib(log.timestamp)
            logs_data.append({
                'id': log.id,
                'timestamp': log_time_wib.isoformat(),
                'status': log.status,
                'is_liveness_real': log.is_liveness_real,
                'failure_count': log.failure_count,
                'message': 'Verifikasi berhasil' if log.status else 'Verifikasi gagal',
                'foto_url': log.foto.url if log.foto else None
            })
        
        return JsonResponse({
            'success': True,
            'has_active_session': True,
            'session': session_data,
            'logs': logs_data,
            'config': {
                'verification_interval': VERIFICATION_INTERVAL,
                'max_consecutive_failures': 2,
                'grace_period_minutes': GRACE_PERIOD_MINUTES
            }
        })
        
    except Mahasiswa.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Mahasiswa tidak ditemukan'
        })
    except Exception as e:
        print(f"Error in get_session_status: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': str(e)
        })
        
@csrf_exempt
@login_required
def auto_checkout_timeout(request):
    """
    API untuk auto checkout ketika mahasiswa tidak memulai monitoring dalam 10 menit
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Method not allowed'})
    
    try:
        data = json.loads(request.body)
        presensi_id = data.get('presensi_id')
        reason = data.get('reason', 'timeout_no_monitoring')
        
        if not presensi_id:
            return JsonResponse({'success': False, 'message': 'presensi_id required'})
        
        mahasiswa = get_object_or_404(Mahasiswa, user=request.user)
        
        presensi = get_object_or_404(
            Presensi,
            id=presensi_id,
            mahasiswa=mahasiswa,
            jam_checkout__isnull=True
        )
        
        now_local = datetime.now()
        
        # Lakukan auto checkout
        presensi.jam_checkout = now_local.time()
        presensi.session_status = 'auto_checkout'
        presensi.consecutive_failures = 0
        
        # Hitung durasi
        checkin_dt = datetime.combine(presensi.tanggal_presensi, presensi.jam_checkin)
        checkout_dt = datetime.combine(presensi.tanggal_presensi, presensi.jam_checkout)
        
        if checkout_dt < checkin_dt:
            checkout_dt += timedelta(days=1)
        
        durasi = checkout_dt - checkin_dt
        
        Durasi.objects.update_or_create(
            presensi=presensi,
            defaults={'waktu_durasi': durasi}
        )
        
        presensi.save()
        
        print(f"[AUTO CHECKOUT] Mahasiswa {mahasiswa.user.nama_lengkap} - {reason}")
        
        return JsonResponse({
            'success': True,
            'message': f'Auto checkout: {reason}',
            'checkout_time': presensi.jam_checkout.strftime('%H:%M:%S')
        })
        
    except Exception as e:
        print(f"Error in auto_checkout_timeout: {e}")
        return JsonResponse({'success': False, 'message': str(e)})

@login_required
def get_verification_logs(request):
    """API untuk mengambil logs verifikasi"""
    try:
        presensi_id = request.GET.get('presensi_id')
        if not presensi_id:
            return JsonResponse({'success': False, 'error': 'presensi_id required'})
        
        # Ambil logs untuk presensi ini
        logs = VerificationLog.objects.filter(
            presensi_id=presensi_id
        ).order_by('-timestamp')[:20]  # Ambil 20 log terakhir
        
        log_data = []
        for log in logs:
            log_data.append({
                'id': log.id,
                'timestamp': log.timestamp.isoformat(),
                'status': log.status,
                'is_liveness_real': log.is_liveness_real,
                'failure_count': log.failure_count,
                'message': 'Verifikasi berhasil' if log.status else 'Verifikasi gagal',
                'foto_url': log.foto.url if log.foto else None
            })
        
        return JsonResponse({
            'success': True,
            'logs': log_data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

def calculate_aggregate_progress(mahasiswa):
    """
    Menghitung progress AGREGAT (total gabungan) dari semua kegiatan PA
    """
    # 1. Hitung total target dari semua kegiatan PA yang diambil
    kegiatan_pa_list = mahasiswa.kegiatan_pa.all()
    
    total_target_jam = 0
    total_sks = 0
    
    for kegiatan in kegiatan_pa_list:
        total_target_jam += kegiatan.target_jam
        total_sks += kegiatan.jumlah_sks
    
    # 2. Hitung total durasi dari SEMUA presensi (AGREGAT - tanpa filter kegiatan_pa)
    presensi_list = Presensi.objects.filter(
        mahasiswa=mahasiswa,
        jam_checkin__isnull=False,
        jam_checkout__isnull=False
    )
    
    total_durasi_detik = 0
    
    for presensi in presensi_list:
        if presensi.jam_checkin and presensi.jam_checkout:
            try:
                # Buat datetime untuk checkin dan checkout
                checkin_dt = datetime.combine(presensi.tanggal_presensi, presensi.jam_checkin)
                checkout_dt = datetime.combine(presensi.tanggal_presensi, presensi.jam_checkout)
                
                # Jika checkout lebih kecil dari checkin (lewat tengah malam)
                if checkout_dt < checkin_dt:
                    checkout_dt += timedelta(days=1)
                
                # Hitung selisih dalam detik
                durasi = checkout_dt - checkin_dt
                total_durasi_detik += durasi.total_seconds()
                
                # Debug log
                print(f"DEBUG - Presensi {presensi.id}: {durasi.total_seconds()} detik")
                
            except Exception as e:
                print(f"Error menghitung durasi presensi {presensi.id}: {e}")
    
    # Konversi detik ke jam
    total_durasi_jam = total_durasi_detik / 3600
    
    # Bulatkan ke 1 angka desimal
    total_durasi_jam = round(total_durasi_jam, 1)
    
    # 3. Hitung progress persentase
    progress_percentage = 0
    if total_target_jam > 0:
        progress_percentage = (total_durasi_jam / total_target_jam) * 100
    
    # Bulatkan progress ke 1 angka desimal
    progress_percentage = round(progress_percentage, 1)
    
    # 4. Debug info
    print(f"\n=== DEBUG AGGREGATE PROGRESS ===")
    print(f"Mahasiswa: {mahasiswa.user.nama_lengkap}")
    print(f"Jumlah Kegiatan: {len(kegiatan_pa_list)}")
    print(f"Total Target Jam: {total_target_jam} jam")
    print(f"Total Durasi: {total_durasi_jam} jam ({total_durasi_detik} detik)")
    print(f"Progress: {progress_percentage}%")
    print(f"Presensi ditemukan: {presensi_list.count()}")
    print(f"===============================\n")
    
    return {
        'total_target_jam': total_target_jam,
        'total_durasi_jam': total_durasi_jam,
        'total_sks': total_sks,
        'progress_percentage': progress_percentage,
        'sisa_jam': max(0, total_target_jam - total_durasi_jam),
        'kegiatan_count': len(kegiatan_pa_list),
        'detail_kegiatan': [
            {
                'nama': kegiatan.nama_kegiatan,
                'sks': kegiatan.jumlah_sks,
                'target_jam': kegiatan.target_jam,
                'jumlah_presensi': Presensi.objects.filter(
                    mahasiswa=mahasiswa,
                    kegiatan_pa=kegiatan,
                    jam_checkin__isnull=False,
                    jam_checkout__isnull=False
                ).count()
            }
            for kegiatan in kegiatan_pa_list
        ]
    }
# Modifikasi fungsi checkout_presensi untuk langsung update jam_tercapai
@csrf_exempt
@login_required
def checkout_presensi(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            mahasiswa_id = data.get('mahasiswa_id')
            foto_base64 = data.get('foto')
            
            # Decode base64 image
            if foto_base64:
                format, imgstr = foto_base64.split(';base64,')
                ext = format.split('/')[-1]
                
                # Buat nama file
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f'checkout_{mahasiswa_id}_{timestamp}.{ext}'
                
                # Simpan foto check-out
                foto_data = ContentFile(base64.b64decode(imgstr), name=filename)
            
            # Get mahasiswa
            mahasiswa = get_object_or_404(Mahasiswa, id=mahasiswa_id)
            
            # Cari presensi hari ini yang belum check-out
            presensi = Presensi.objects.filter(
                mahasiswa=mahasiswa,
                tanggal_presensi=date.today(),
                jam_checkout__isnull=True
            ).order_by('-jam_checkin').first()
            
            if not presensi:
                return JsonResponse({
                    'success': False,
                    'message': 'Tidak ada sesi check-in yang aktif'
                })
            
            # Update presensi dengan check-out
            presensi.jam_checkout = datetime.now().time()
            
            # Simpan foto jika ada
            if foto_base64:
                presensi.foto_checkout = foto_data
                
            presensi.save()
            
            # Hitung durasi - FIX INI: Gunakan datetime yang benar
            if presensi.jam_checkin and presensi.jam_checkout:
                checkin_dt = datetime.combine(
                    presensi.tanggal_presensi, 
                    presensi.jam_checkin
                )
                checkout_dt = datetime.combine(
                    presensi.tanggal_presensi, 
                    presensi.jam_checkout
                )
                
                # Jika checkout lebih kecil dari checkin (lewat tengah malam)
                if checkout_dt < checkin_dt:
                    checkout_dt += timedelta(days=1)
                
                durasi = checkout_dt - checkin_dt
                
                # Debug: Cetak durasi
                print(f"DEBUG checkout_presensi - Durasi dihitung: {durasi}")
                print(f"DEBUG - Checkin: {checkin_dt}, Checkout: {checkout_dt}")
                
                # Simpan durasi - PERBAIKAN: Buat atau update Durasi
                Durasi.objects.update_or_create(
                    presensi=presensi,
                    defaults={'waktu_durasi': durasi}
                )
                
                print(f"DEBUG - Durasi berhasil disimpan untuk presensi {presensi.id}")
                
                # PERBAIKAN: Hitung total durasi untuk update Status_Pemenuhan_SKS
                total_hours = calculate_total_duration(mahasiswa.id, presensi.kegiatan_pa.id)
                
                # Debug
                print(f"DEBUG - Total hours setelah checkout: {total_hours}")
                
                # Update Status_Pemenuhan_SKS
                status_sks, created = Status_Pemenuhan_SKS.objects.get_or_create(
                    mahasiswa=mahasiswa,
                    kegiatan_pa=presensi.kegiatan_pa,
                    defaults={
                        'jam_target': presensi.kegiatan_pa.target_jam,
                        'jumlah_sks': presensi.kegiatan_pa.jumlah_sks,
                        'jam_tercapai': total_hours
                    }
                )
                
                if not created:
                    status_sks.jam_tercapai = total_hours
                
                # Cek apakah sudah memenuhi target
                if status_sks.jam_tercapai >= status_sks.jam_target:
                    status_sks.status_pemenuhan = 'memenuhi'
                else:
                    status_sks.status_pemenuhan = 'belum memenuhi'
                
                status_sks.save()
                
                print(f"DEBUG - Status SKS updated: {status_sks.jam_tercapai} jam tercapai")
            
            # Simpan juga ke FotoWajah untuk dataset
            if foto_base64:
                FotoWajah.objects.create(
                    mahasiswa=mahasiswa,
                    file_path=foto_data,
                    keterangan=f'Check-out {datetime.now().strftime("%d/%m/%Y %H:%M")}'
                )
            
            return JsonResponse({
                'success': True,
                'message': 'Check-out berhasil',
                'data': {
                    'jam_checkout': presensi.jam_checkout.strftime('%H:%M'),
                    'presensi_id': presensi.id,
                    'total_durasi_jam': total_hours if 'total_hours' in locals() else 0
                }
            })
            
        except Exception as e:
            import traceback
            return JsonResponse({
                'success': False,
                'message': f'Error: {str(e)}',
                'trace': traceback.format_exc()
            })
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'})

@login_required
def fix_missing_durations(request):
    """
    Fungsi untuk membuat Durasi dari presensi yang sudah ada
    tapi tidak memiliki Durasi (hanya memperbaiki tabel Durasi saja).
    Tidak lagi update Status_Pemenuhan_SKS karena sistem sekarang agregat.
    """
    if not (request.user.role == 'admin' or request.user.is_superuser):
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    try:
        # Cari semua presensi yang sudah check-in & check-out tapi belum punya Durasi
        presensi_without_duration = Presensi.objects.filter(
            jam_checkin__isnull=False,
            jam_checkout__isnull=False
        ).exclude(
            durasi__isnull=False  # Exclude yang sudah punya Durasi
        )
        
        fixed_count = 0
        
        for presensi in presensi_without_duration:
            # Hitung durasi
            checkin_dt = datetime.combine(presensi.tanggal_presensi, presensi.jam_checkin)
            checkout_dt = datetime.combine(presensi.tanggal_presensi, presensi.jam_checkout)
            
            if checkout_dt < checkin_dt:
                checkout_dt += timedelta(days=1)
            
            durasi = checkout_dt - checkin_dt
            
            # Buat record Durasi
            Durasi.objects.create(
                presensi=presensi,
                waktu_durasi=durasi
            )
            
            print(f"Fixed: Presensi {presensi.id} - Durasi: {durasi}")
            fixed_count += 1
        
        return JsonResponse({
            'success': True,
            'message': f'Berhasil memperbaiki {fixed_count} data Durasi yang hilang',
            'fixed_count': fixed_count
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}'
        })

@login_required
def check_duration_status(request):
    """
    Cek status Durasi untuk debugging
    """
    try:
        mahasiswa = Mahasiswa.objects.get(user=request.user)
        
        # Cek semua presensi mahasiswa
        presensi_list = Presensi.objects.filter(
            mahasiswa=mahasiswa
        ).select_related('kegiatan_pa')
        
        result = []
        
        for presensi in presensi_list:
            has_duration = hasattr(presensi, 'durasi')
            durasi_value = presensi.durasi.waktu_durasi if has_duration else None
            
            result.append({
                'id': presensi.id,
                'tanggal': presensi.tanggal_presensi,
                'kegiatan': presensi.kegiatan_pa.nama_kegiatan,
                'checkin': presensi.jam_checkin,
                'checkout': presensi.jam_checkout,
                'has_duration': has_duration,
                'duration_value': str(durasi_value) if durasi_value else None,
            })
        
        # Debug info
        total_presensi = len(result)
        presensi_with_duration = sum(1 for r in result if r['has_duration'])
        
        print(f"DEBUG check_duration_status - Total presensi: {total_presensi}")
        print(f"DEBUG - Presensi dengan Durasi: {presensi_with_duration}")
        print(f"DEBUG - Presensi tanpa Durasi: {total_presensi - presensi_with_duration}")
        
        return JsonResponse({
            'success': True,
            'data': result,
            'summary': {
                'total_presensi': total_presensi,
                'with_duration': presensi_with_duration,
                'without_duration': total_presensi - presensi_with_duration
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}'
        })

@login_required
def get_progress_sks_api(request):
    """API untuk mendapatkan progress SKS berdasarkan durasi"""
    try:
        user = request.user
        mahasiswa = Mahasiswa.objects.get(user=user)
        
        # Ambil semua status pemenuhan SKS untuk mahasiswa ini (hanya kegiatan yang diambil)
        status_list = Status_Pemenuhan_SKS.objects.filter(
            mahasiswa=mahasiswa,
            kegiatan_pa__in=mahasiswa.kegiatan_pa.all()
        ).select_related('kegiatan_pa')
        
        progress_data = []
        total_jam_tercapai = 0
        total_jam_target = 0
        
        for status in status_list:
            # Hitung total durasi untuk kegiatan ini
            total_hours = calculate_total_duration(mahasiswa.id, status.kegiatan_pa.id)
            
            # Update dengan data terbaru
            status.jam_tercapai = total_hours
            if status.jam_tercapai >= status.jam_target:
                status.status_pemenuhan = 'memenuhi'
            else:
                status.status_pemenuhan = 'belum memenuhi'
            status.save()
            
            progress_data.append({
                'kegiatan': status.kegiatan_pa.nama_kegiatan,
                'sks': status.jumlah_sks,
                'jam_target': status.jam_target,
                'jam_tercapai': status.jam_tercapai,
                'progress_percentage': min(100, int((status.jam_tercapai / status.jam_target) * 100)) if status.jam_target > 0 else 0,
                'status': status.status_pemenuhan
            })
            
            total_jam_tercapai += status.jam_tercapai
            total_jam_target += status.jam_target
        
        # Hitung progress keseluruhan
        overall_progress = min(100, int((total_jam_tercapai / total_jam_target) * 100)) if total_jam_target > 0 else 0
        
        return JsonResponse({
            'success': True,
            'data': {
                'mahasiswa': {
                    'nama': mahasiswa.user.nama_lengkap,
                    'nim': mahasiswa.nim,
                    'kelas': mahasiswa.kelas
                },
                'progress_per_kegiatan': progress_data,
                'summary': {
                    'total_jam_tercapai': total_jam_tercapai,
                    'total_jam_target': total_jam_target,
                    'overall_progress_percentage': overall_progress,
                    'sisa_jam': max(0, total_jam_target - total_jam_tercapai)
                }
            }
        })
        
    except Mahasiswa.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Data mahasiswa tidak ditemukan'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}'
        }, status=500)

@login_required
def get_presensi_today(request):
    try:
        today = date.today()
        mahasiswa_list = Mahasiswa.objects.select_related('user').filter(
            pengajuan_pendaftaran__status_pengajuan='disetujui'
        )  # tambah filter approved biar aman

        presensi_list = []
        for mahasiswa in mahasiswa_list:
            presensi_today = Presensi.objects.filter(
                mahasiswa=mahasiswa,
                tanggal_presensi=today
            ).order_by('-jam_checkin')

            presensi_data = []
            for presensi in presensi_today:
                presensi_data.append({
                    'jam_checkin': presensi.jam_checkin.strftime('%H:%M') if presensi.jam_checkin else '-',
                    'jam_checkout': presensi.jam_checkout.strftime('%H:%M') if presensi.jam_checkout else '-',
                    'foto_checkin_url': presensi.foto_checkin.url if presensi.foto_checkin else '',
                    'foto_checkout_url': presensi.foto_checkout.url if presensi.foto_checkout else ''
                })

            is_checked_in = any(p.jam_checkin and not p.jam_checkout for p in presensi_today)

            presensi_list.append({
                'id': mahasiswa.id,
                'nama': mahasiswa.user.nama_lengkap,
                'nrp': mahasiswa.user.nrp,
                'kelas': mahasiswa.kelas,
                'is_checked_in': is_checked_in,
                'presensi_today': presensi_data
            })

        return JsonResponse({
            'success': True,
            'data': presensi_list,
            'today': today.strftime('%d/%m/%Y')
        })

    except Exception as e:
        import traceback
        print(f"[ERROR get_presensi_today] {str(e)}")
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
        
@login_required
def profil_mahasiswa(request):
    user = request.user
    try:
        mahasiswa = Mahasiswa.objects.select_related(
            'jenjang_pendidikan', 
            'user',
            'semester',
            'kelas'
        ).prefetch_related('kegiatan_pa').get(user=user)
        
        nama_lengkap = user.nama_lengkap or user.get_full_name() or user.username
        nrp = user.nrp or user.username
        
        kegiatan_pa_selected = list(mahasiswa.kegiatan_pa.all())
        import json
        kegiatan_ids = [k.id for k in kegiatan_pa_selected]
        
        # Get dosen pembimbing
        dosen_pembimbing_qs = Mahasiswa_Dosen.objects.filter(
            mahasiswa=mahasiswa
        ).select_related('dosen').order_by('tipe_pembimbing')
        
        dosen_pembimbing = []
        for idx, md in enumerate(dosen_pembimbing_qs, 1):
            dosen_pembimbing.append({
                'nomor': idx,
                'dosen': md.dosen,
                'nama_dosen': md.dosen.nama_dosen,
                'nip': md.dosen.nip,
                'prodi': md.dosen.prodi,
            })
        
        # Get all data for dropdowns
        dosen_list = Dosen.objects.all().order_by('nama_dosen')
        jenjang_list = Jenjang_Pendidikan.objects.all()
        kelas_list = Kelas.objects.filter(is_active=True).order_by('nama_kelas')
        
        # PERBAIKAN: Siapkan semester_list_json untuk initial load
        semester_list_json = []
        if mahasiswa.jenjang_pendidikan:
            jenjang_nama = mahasiswa.jenjang_pendidikan.nama_jenjang
            if "D3" in jenjang_nama:
                semesters = Semester.objects.filter(jenjang='D3').order_by('nomor_semester')
            elif "D4" in jenjang_nama:
                semesters = Semester.objects.filter(jenjang='D4').order_by('nomor_semester')
            elif "S2" in jenjang_nama:
                semesters = Semester.objects.filter(jenjang='S2').order_by('nomor_semester')
            else:
                semesters = []
            
            semester_list_json = [{'id': s.id, 'nama_semester': s.nama_semester} for s in semesters]
        
        context = {
            'mahasiswa': {
                'nama': nama_lengkap,
                'nrp': nrp,
                'email': user.email,
                'jenjang': mahasiswa.jenjang_pendidikan.nama_jenjang if mahasiswa.jenjang_pendidikan else '',
                'jenjang_pendidikan': mahasiswa.jenjang_pendidikan,
                'kelas': mahasiswa.kelas.nama_kelas if mahasiswa.kelas else '',
                'kelas_obj': mahasiswa.kelas,
                'semester': mahasiswa.semester.nama_semester if mahasiswa.semester else '',
                'semester_id': mahasiswa.semester.id if mahasiswa.semester else '',
                'kegiatan_pa': kegiatan_pa_selected,
            },
            'dosen_pembimbing': dosen_pembimbing,
            'dosen_list': dosen_list,
            'jenjang_list': jenjang_list,
            'kelas_list': kelas_list,
            'semester_list_json': json.dumps(semester_list_json),  # Tambahkan ini!
            'kegiatan_ids_json': json.dumps(kegiatan_ids),
        }
        
    except Mahasiswa.DoesNotExist as e:
        print(f"Error: Mahasiswa not found for user {user.username}: {e}")
        context = {
            'mahasiswa': None,
            'dosen_pembimbing': [],
            'dosen_list': [],
            'jenjang_list': [],
            'kelas_list': [],
            'semester_list_json': '[]',
            'kegiatan_ids_json': '[]',
        }
        
    return render(request, 'mahasiswa/profil_mahasiswa.html', context)

@login_required
def edit_dosen_pembimbing(request):
    if request.method == 'POST':
        try:
            user = request.user
            mahasiswa = Mahasiswa.objects.get(user=user)
            
            # Validasi dosen pembimbing
            dosen1_id = request.POST.get('dosen_pembimbing1')
            dosen2_id = request.POST.get('dosen_pembimbing2')
            dosen3_id = request.POST.get('dosen_pembimbing3')
            
            # Validasi required fields
            if not dosen1_id or not dosen2_id:
                return JsonResponse({'success': False, 'error': 'Dosen pembimbing 1 dan 2 wajib diisi'})
            
            # Validasi dosen tidak sama
            if dosen1_id == dosen2_id:
                return JsonResponse({'success': False, 'error': 'Dosen pembimbing 1 dan 2 tidak boleh sama'})
            
            if dosen3_id and (dosen3_id == dosen1_id or dosen3_id == dosen2_id):
                return JsonResponse({'success': False, 'error': 'Dosen pembimbing 3 tidak boleh sama dengan dosen lainnya'})
            
            # Delete existing dosen pembimbing
            Mahasiswa_Dosen.objects.filter(mahasiswa=mahasiswa).delete()
            
            # Add new dosen pembimbing
            dosen1 = Dosen.objects.get(id=dosen1_id)
            Mahasiswa_Dosen.objects.create(
                mahasiswa=mahasiswa, 
                dosen=dosen1, 
                tipe_pembimbing='pembimbing1'
            )
            
            dosen2 = Dosen.objects.get(id=dosen2_id)
            Mahasiswa_Dosen.objects.create(
                mahasiswa=mahasiswa, 
                dosen=dosen2, 
                tipe_pembimbing='pembimbing2'
            )
            
            if dosen3_id:
                dosen3 = Dosen.objects.get(id=dosen3_id)
                Mahasiswa_Dosen.objects.create(
                    mahasiswa=mahasiswa, 
                    dosen=dosen3, 
                    tipe_pembimbing='pembimbing3'
                )
            
            return JsonResponse({'success': True, 'message': 'Dosen pembimbing berhasil diperbarui'})
            
        except Dosen.DoesNotExist as e:
            return JsonResponse({'success': False, 'error': f'Dosen tidak ditemukan: {str(e)}'})
        except Exception as e:
            print(f"Error saving dosen: {e}")  # Untuk debugging
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

def get_kegiatan_pa_by_jenjang(request, jenjang_id):
    """API untuk mendapatkan kegiatan PA berdasarkan jenjang"""
    try:
        kegiatan = Kegiatan_PA.objects.filter(jenjang_pendidikan_id=jenjang_id)
        kegiatan_list = list(kegiatan.values('id', 'nama_kegiatan', 'jumlah_sks', 'target_jam'))
        return JsonResponse({
            'success': True,
            'kegiatan': kegiatan_list
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)
    
def upload_foto_wajah(request):
    if request.method == 'POST' and request.FILES.get('foto_wajah'):
        try:
            user = request.user
            mahasiswa = Mahasiswa.objects.get(user=user)
            
            foto = request.FILES['foto_wajah']
            keterangan = request.POST.get('keterangan', '')
            
            # Simpan foto
            foto_wajah = FotoWajah.objects.create(
                mahasiswa=mahasiswa,
                file_path=foto,
                keterangan=keterangan
            )
            
            return JsonResponse({
                'success': True,
                'message': 'Foto berhasil diupload',
                'foto_id': foto_wajah.id
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    return JsonResponse({'success': False, 'error': 'Invalid request'})

@login_required
def progress_sks(request):
    try:
        mahasiswa = Mahasiswa.objects.get(user=request.user)
        
        # Hitung progress AGREGAT
        progress_data = calculate_aggregate_progress(mahasiswa)
        
        total_target = progress_data['total_target_jam']
        total_tercapai = progress_data['total_durasi_jam']
        total_sks = progress_data['total_sks']
        progress_percentage = progress_data['progress_percentage']
        jam_sisa = progress_data['sisa_jam']
        jumlah_kegiatan = progress_data['kegiatan_count']
        detail_kegiatan = progress_data['detail_kegiatan']
        
        # Hitung statistik lainnya
        rata_per_minggu = 0
        if total_tercapai > 0:
            # Hitung jumlah minggu berjalan sejak semester dimulai
            try:
                if mahasiswa.semester and mahasiswa.semester.tanggal_mulai:
                    tanggal_mulai = mahasiswa.semester.tanggal_mulai
                    hari_berjalan = (date.today() - tanggal_mulai).days
                    minggu_berjalan = max(1, math.ceil(hari_berjalan / 7))
                else:
                    minggu_berjalan = 8  # Default 8 minggu
                
                rata_per_minggu = round(total_tercapai / minggu_berjalan, 1)
            except:
                rata_per_minggu = round(total_tercapai / 8, 1)  # Fallback
        
        estimasi_selesai = ""
        sisa_waktu = 0
        rekomendasi_per_hari = 0
        
        if rata_per_minggu > 0 and jam_sisa > 0:
            sisa_minggu = jam_sisa / rata_per_minggu
            tanggal_estimasi = date.today() + timedelta(days=sisa_minggu * 7)
            estimasi_selesai = tanggal_estimasi.strftime("%d %B %Y")
            sisa_waktu = int(sisa_minggu * 7)
            rekomendasi_per_hari = round(jam_sisa / sisa_waktu, 2) if sisa_waktu > 0 else 0
        
        # Data untuk JavaScript
        js_data = {
            'progress_percentage': progress_percentage,
            'jam_terselesaikan': total_tercapai,
            'jam_sisa': jam_sisa,
            'total_sks': total_sks,
            'jam_target_total': total_target,
            'rata_per_minggu': rata_per_minggu,
            'sisa_waktu': sisa_waktu,
            'rekomendasi_per_hari': rekomendasi_per_hari,
            'estimasi_selesai': estimasi_selesai,
            'mahasiswa_nama': mahasiswa.user.nama_lengkap,
            'mahasiswa_nim': mahasiswa.nim,
            'semester_nama': mahasiswa.semester.nama_semester if mahasiswa.semester else "-",
            'kegiatan_count': jumlah_kegiatan,
        }
        
        # Debug info ke console
        print(f"\n=== DEBUG PROGRESS SKS PAGE ===")
        print(f"Mahasiswa: {mahasiswa.user.nama_lengkap} ({mahasiswa.nim})")
        print(f"Total Presensi: {Presensi.objects.filter(mahasiswa=mahasiswa).count()}")
        print(f"Total Durasi: {total_tercapai} jam")
        print(f"Target: {total_target} jam")
        print(f"Progress: {progress_percentage}%")
        print("==============================\n")
        
        context = {
            'mahasiswa': mahasiswa,
            'total_sks': total_sks,
            'jam_terselesaikan': total_tercapai,
            'jam_target_total': total_target,
            'jam_sisa': jam_sisa,
            'progress_percentage': progress_percentage,
            'rata_per_minggu': rata_per_minggu,
            'estimasi_selesai': estimasi_selesai,
            'sisa_waktu': sisa_waktu,
            'rekomendasi_per_hari': rekomendasi_per_hari,
            'jumlah_kegiatan': jumlah_kegiatan,
            'detail_kegiatan': detail_kegiatan,
            'js_data_json': mark_safe(json.dumps(js_data)),
        }
        
        return render(request, 'mahasiswa/progress_sks.html', context)
        
    except Mahasiswa.DoesNotExist:
        messages.error(request, 'Data mahasiswa tidak ditemukan')
        return redirect('profil_mahasiswa')
    except Exception as e:
        print(f"Error in progress_sks: {e}")
        import traceback
        traceback.print_exc()
        messages.error(request, f'Terjadi kesalahan saat mengambil data progress SKS: {str(e)}')
        return redirect('profil_mahasiswa')

# --- ADMIN VIEWS ---
@login_required
def admin_dashboard(request):
    if not (request.user.role == 'admin' or request.user.is_superuser):
        return redirect('login')

    today = date.today()
    week_start = today - timedelta(days=today.weekday())  # Senin minggu ini

    # Nama hari yang diterjemahkan (lazy)
    DAY_NAMES = [
        _("Senin"), _("Selasa"), _("Rabu"), _("Kamis"),
        _("Jumat"), _("Sabtu"), _("Minggu")
    ]
    day_labels = {i: str(DAY_NAMES[i]) for i in range(7)}

    # Presensi hari ini
    today_present = Presensi.objects.filter(
        tanggal_presensi=today, jam_checkin__isnull=False
    ).values('mahasiswa_id').distinct().count()

    total_approved_students = Mahasiswa.objects.filter(
        pengajuan_pendaftaran__status_pengajuan='disetujui'
    ).count()

    # Mapping warna & badge
    color_map = {
        'D3 - Diploma 3': 'bg-primary',
        'D4 - Diploma 4': 'bg-info',
        'LJ - Lanjut Jenjang': 'bg-success',
        'S2 - Magister': 'bg-warning text-dark',
    }
    badge_map = {
        'D3 - Diploma 3': 'primary',
        'D4 - Diploma 4': 'info',
        'LJ - Lanjut Jenjang': 'success',
        'S2 - Magister': 'warning',
    }

    jenjang_list = Jenjang_Pendidikan.objects.all().order_by('nama_jenjang')

    weekly_data = {}  # key: 0-6 (Senin-Minggu)
    weekly_summary = []
    total_hadir_week = total_mahasiswa_week = 0

    # Data per hari dalam seminggu
    for i in range(7):
        day_date = week_start + timedelta(days=i)
        weekday_num = day_date.weekday()

        # Max hadir hari itu (untuk scaling bar)
        max_hadir_hari = Presensi.objects.filter(
            tanggal_presensi=day_date,
            jam_checkin__isnull=False
        ).values('mahasiswa_id').distinct().count() or 1

        day_entries = []
        for jenjang in jenjang_list:
            hadir = Presensi.objects.filter(
                tanggal_presensi=day_date,
                mahasiswa__jenjang_pendidikan=jenjang,
                jam_checkin__isnull=False
            ).values('mahasiswa_id').distinct().count()

            total_mhs = Mahasiswa.objects.filter(
                jenjang_pendidikan=jenjang,
                pengajuan_pendaftaran__status_pengajuan='disetujui'
            ).count()

            perc = round(hadir / total_mhs * 100, 1) if total_mhs > 0 else 0.0

            # Scaling height: lebih aman, hindari height 0% meski ada hadir
            height_scaled = 0
            if max_hadir_hari > 0 and hadir > 0:
                height_scaled = round((hadir / max_hadir_hari) * 100, 1)
            # Opsional: minimal height 5% kalau ada hadir tapi scaling kecil
            if hadir > 0 and height_scaled < 5:
                height_scaled = 5.0

            day_entries.append({
                'jenjang': jenjang.nama_jenjang,
                'hadir': hadir,
                'total': total_mhs,
                'percentage': perc,
                'height': height_scaled,
                'color': color_map.get(jenjang.nama_jenjang, 'bg-primary'),
                'badge': badge_map.get(jenjang.nama_jenjang, 'primary'),
            })

        weekly_data[weekday_num] = day_entries

    # Ringkasan mingguan per jenjang
    for jenjang in jenjang_list:
        hadir_week = Presensi.objects.filter(
            tanggal_presensi__gte=week_start,
            tanggal_presensi__lte=today,
            mahasiswa__jenjang_pendidikan=jenjang,
            jam_checkin__isnull=False
        ).values('mahasiswa_id').distinct().count()

        total_mhs = Mahasiswa.objects.filter(
            jenjang_pendidikan=jenjang,
            pengajuan_pendaftaran__status_pengajuan='disetujui'
        ).count()

        perc_week = round(hadir_week / total_mhs * 100, 1) if total_mhs > 0 else 0.0

        weekly_summary.append({
            'code': jenjang.nama_jenjang,
            'present': hadir_week,
            'total': total_mhs,
            'percentage': perc_week,
            'color': color_map.get(jenjang.nama_jenjang, 'bg-primary'),
            'badge': badge_map.get(jenjang.nama_jenjang, 'primary'),
        })

        total_hadir_week += hadir_week
        total_mahasiswa_week += total_mhs

    weekly_total_perc = round(total_hadir_week / total_mahasiswa_week * 100, 1) if total_mahasiswa_week > 0 else 0.0

    # Rata-rata mingguan
    daily_percentages = []
    for entries in weekly_data.values():
        day_hadir = sum(e['hadir'] for e in entries)
        day_total = sum(e['total'] for e in entries)
        daily_percentages.append(round(day_hadir / day_total * 100, 1) if day_total > 0 else 0.0)
    avg_attendance_week = round(sum(daily_percentages) / len(daily_percentages), 1) if daily_percentages else 0.0

    # Semester aktif
    try:
        active_ta = Tahun_Ajaran.objects.get(status_aktif='aktif')
        current_semester_name = active_ta.nama_tahun_ajaran or "Semester Aktif"
        days_total = (active_ta.tanggal_selesai - active_ta.tanggal_mulai).days
        days_passed = (today - active_ta.tanggal_mulai).days
        semester_progress = round(days_passed / days_total * 100) if days_total > 0 else 0
        weeks_remaining = max(0, round((days_total - days_passed) / 7))
    except Tahun_Ajaran.DoesNotExist:
        current_semester_name = "Tidak Ada Semester Aktif"
        semester_progress = weeks_remaining = 0

    # Recent activities
    recent_presensi = Presensi.objects.select_related('mahasiswa__user').filter(
        jam_checkin__isnull=False
    ).order_by('-tanggal_presensi', '-jam_checkin')[:4]

    recent_activities = []
    for p in recent_presensi:
        dt = timezone.localtime(timezone.make_aware(datetime.combine(p.tanggal_presensi, p.jam_checkin)))
        time_ago = dt.strftime("%H:%M") if p.tanggal_presensi == today else p.tanggal_presensi.strftime("%d %b")
        initials = ''.join([w[0].upper() for w in (p.mahasiswa.user.nama_lengkap or "").split()[:2]]) or "??"

        recent_activities.append({
            'nama': p.mahasiswa.user.nama_lengkap or p.mahasiswa.nim,
            'nim': p.mahasiswa.nim,
            'action': "Check-out" if p.jam_checkout else "Check-in",
            'time_ago': time_ago,
            'color': 'success' if p.jam_checkout else 'primary',
            'initials': initials,
        })

    # JSON untuk modal
    day_data_json = {}
    for weekday_num, entries in weekly_data.items():
        day_data_json[str(weekday_num)] = [
            {
                'jenjang': e['jenjang'],
                'hadir': e['hadir'],
                'total': e['total'],
                'percentage': e['percentage'],
                'badge': e['badge']
            } for e in entries
        ]

    # Siapkan context
    context = {
        'today_present_count': today_present,
        'today_total_students': total_approved_students,
        'avg_attendance_week': avg_attendance_week,
        'current_semester_name': current_semester_name,
        'semester_progress': semester_progress,
        'weeks_remaining': weeks_remaining,
        'weekly_data': weekly_data,
        'weekly_summary': weekly_summary,
        'weekly_total': {
            'present': total_hadir_week,
            'total': total_mahasiswa_week,
            'percentage': weekly_total_perc,
        },
        'recent_activities': recent_activities,
        'day_data_json': day_data_json,
        'day_labels': day_labels,
    }

    # Pastikan semua hari ada (untuk chart lengkap)
    for i in range(7):
        if i not in weekly_data:
            weekly_data[i] = []

    # Sorted untuk urutan Senin-Minggu
    sorted_weekly = sorted(weekly_data.items(), key=lambda x: x[0])
    context['sorted_weekly'] = sorted_weekly

    return render(request, 'admin/admin_dashboard.html', context)

@login_required
def monitoring_presensi(request):
    """View untuk monitoring presensi admin"""
    return render(request, 'admin/status_pemenuhan_sks.html')

@login_required
def status_pemenuhan_sks(request):
    if not (request.user.role == 'admin' or request.user.is_superuser):
        messages.error(request, 'Akses ditolak. Halaman ini untuk admin.')
        return redirect('admin_dashboard')

    mahasiswa_setuju = Mahasiswa.objects.filter(
        pengajuan_pendaftaran__status_pengajuan='disetujui'
    ).select_related('user', 'jenjang_pendidikan').prefetch_related('kegiatan_pa').distinct()

    jenjang_list = Jenjang_Pendidikan.objects.all().order_by('nama_jenjang')
    data_mahasiswa = []

    for mahasiswa in mahasiswa_setuju:
        # Hitung TOTAL agregat (sama seperti halaman mahasiswa)
        total_jam_ditempuh = calculate_total_duration_all(mahasiswa.id)

        kegiatan_pa_set = mahasiswa.kegiatan_pa.all()
        total_jam_target = sum(k.target_jam for k in kegiatan_pa_set)
        total_sks = sum(k.jumlah_sks for k in kegiatan_pa_set)

        status_overall = 'Memenuhi' if total_jam_ditempuh >= total_jam_target else 'Belum Memenuhi'

        progress_percentage = 0
        progress_class = 'bg-secondary'
        if total_jam_target > 0:
            progress_percentage = (total_jam_ditempuh / total_jam_target) * 100
            if total_jam_ditempuh >= total_jam_target:
                progress_class = 'bg-success'
            elif total_jam_ditempuh >= total_jam_target * 0.7:
                progress_class = 'bg-primary'
            elif total_jam_ditempuh >= total_jam_target * 0.4:
                progress_class = 'bg-warning'
            else:
                progress_class = 'bg-danger'

        sisa_jam = max(0, total_jam_target - total_jam_ditempuh)
        persentase_40 = total_jam_target * 0.4
        persentase_70 = total_jam_target * 0.7

        # Breakdown per kegiatan (proporsional untuk tampilan saja)
        kegiatan_list = []
        for kegiatan in kegiatan_pa_set:
            proporsi = kegiatan.target_jam / total_jam_target if total_jam_target > 0 else 0
            jam_tercapai_proporsional = round(total_jam_ditempuh * proporsi, 1)

            kegiatan_list.append({
                'nama_kegiatan': kegiatan.nama_kegiatan,
                'sks': kegiatan.jumlah_sks,
                'jam_per_minggu': getattr(kegiatan, 'total_jam_minggu', 0),
                'jam_target': kegiatan.target_jam,
                'jam_tercapai': jam_tercapai_proporsional,
                'status_per_kegiatan': 'Memenuhi' if jam_tercapai_proporsional >= kegiatan.target_jam else 'Belum Memenuhi'
            })

        data_mahasiswa.append({
            'nama': mahasiswa.user.nama_lengkap or mahasiswa.user.username,
            'jenjang': mahasiswa.jenjang_pendidikan.nama_jenjang if mahasiswa.jenjang_pendidikan else '-',
            'kegiatan_list': kegiatan_list,
            'total_jam_ditempuh': round(total_jam_ditempuh, 1),
            'total_jam_target': round(total_jam_target, 1),
            'total_sks': total_sks,
            'progress_percentage': round(progress_percentage, 1),
            'progress_class': progress_class,
            'sisa_jam': round(sisa_jam, 1),
            'persentase_40': round(persentase_40, 1),
            'persentase_70': round(persentase_70, 1),
            'status_overall': status_overall
        })

    data_mahasiswa.sort(key=lambda x: x['nama'].lower())

    total_memenuhi = sum(1 for m in data_mahasiswa if m['status_overall'] == 'Memenuhi')
    total_belum_memenuhi = sum(1 for m in data_mahasiswa if m['status_overall'] == 'Belum Memenuhi')
    total_sks_all = sum(m['total_sks'] for m in data_mahasiswa)

    context = {
        'data_mahasiswa': data_mahasiswa,
        'total_mahasiswa': len(data_mahasiswa),
        'total_sks_all': total_sks_all,
        'total_memenuhi': total_memenuhi,
        'total_belum_memenuhi': total_belum_memenuhi,
        'jenjang_list': jenjang_list,
    }
    return render(request, 'admin/status_pemenuhan_sks.html', context)

def calculate_total_duration_all(mahasiswa_id):
    """Total jam dari SEMUA presensi mahasiswa, tanpa peduli kegiatan_pa"""
    presensi_list = Presensi.objects.filter(
        mahasiswa_id=mahasiswa_id,
        jam_checkin__isnull=False,
        jam_checkout__isnull=False
    )
    
    total_seconds = 0
    for presensi in presensi_list:
        if presensi.jam_checkin and presensi.jam_checkout:
            checkin_dt = datetime.combine(presensi.tanggal_presensi, presensi.jam_checkin)
            checkout_dt = datetime.combine(presensi.tanggal_presensi, presensi.jam_checkout)
            if checkout_dt < checkin_dt:
                checkout_dt += timedelta(days=1)
            delta = checkout_dt - checkin_dt
            total_seconds += delta.total_seconds()
    
    return round(total_seconds / 3600, 1)

@login_required
def monitor_durasi(request):
    """View untuk monitor durasi presensi mahasiswa"""
    is_checked_in = False
    active_presensi = None
    
    if request.user.role == 'mahasiswa':
        try:
            mahasiswa = Mahasiswa.objects.get(user=request.user)
            # Cari presensi hari ini yang belum checkout
            active_presensi = Presensi.objects.filter(
                mahasiswa=mahasiswa,
                tanggal_presensi=date.today(),
                jam_checkout__isnull=True
            ).order_by('-jam_checkin').first()
            
            is_checked_in = active_presensi is not None
            
        except Mahasiswa.DoesNotExist:
            pass
            
    return render(request, 'mahasiswa/monitor_durasi.html', {
        'is_checked_in': is_checked_in,
        'active_presensi': active_presensi
    })

@login_required
def get_monitoring_status(request):
    """API untuk mendapatkan status monitoring terbaru dari server"""
    try:
        if request.user.role != 'mahasiswa':
            return JsonResponse({'success': False, 'error': 'Bukan mahasiswa'})
        
        mahasiswa = Mahasiswa.objects.get(user=request.user)
        today = date.today()
        
        # Cari presensi aktif
        active_presensi = Presensi.objects.filter(
            mahasiswa=mahasiswa,
            tanggal_presensi=today,
            jam_checkout__isnull=True
        ).first()
        
        if not active_presensi:
            return JsonResponse({
                'success': True,
                'is_checked_in': False
            })
        
        # Hitung sisa waktu sampai verifikasi berikutnya
        now = datetime.now()
        last_verified = active_presensi.last_verified_at or datetime.combine(today, active_presensi.jam_checkin)
        
        # Gunakan timezone-aware jika perlu
        if timezone.is_naive(last_verified):
            last_verified = timezone.make_aware(last_verified)
        if timezone.is_naive(now):
            now = timezone.make_aware(now)
        
        seconds_since_last = (now - last_verified).total_seconds()
        VERIFICATION_INTERVAL = 5 * 60  # 5 menit
        time_left = max(0, VERIFICATION_INTERVAL - seconds_since_last)
        
        # Hitung durasi sejak check-in
        checkin_dt = datetime.combine(today, active_presensi.jam_checkin)
        if timezone.is_naive(checkin_dt):
            checkin_dt = timezone.make_aware(checkin_dt)
        
        duration_seconds = (now - checkin_dt).total_seconds()
        
        return JsonResponse({
            'success': True,
            'is_checked_in': True,
            'presensi_id': active_presensi.id,
            'jam_checkin': active_presensi.jam_checkin.strftime('%H:%M:%S'),
            'tanggal_presensi': active_presensi.tanggal_presensi.isoformat(),
            'time_left': int(time_left),
            'failure_count': active_presensi.failure_count,
            'monitoring_status': active_presensi.monitoring_status,
            'duration_seconds': int(duration_seconds),
            'last_verified_at': active_presensi.last_verified_at.isoformat() if active_presensi.last_verified_at else None
        })
        
    except Mahasiswa.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Mahasiswa tidak ditemukan'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def management_data(request):
    """View untuk management data admin"""
    return render(request, 'admin/management_data.html')

@login_required
def approval_pendaftaran(request):
    """View untuk approval pendaftaran admin"""
    
    search_query = request.GET.get('search', '').strip()
    
    pendaftaran_list = Pengajuan_Pendaftaran.objects.select_related(
        'mahasiswa__user', 
        'mahasiswa__jenjang_pendidikan',
        'mahasiswa__semester'
    ).all()
    
    if search_query:
        pendaftaran_list = pendaftaran_list.filter(
            Q(mahasiswa__user__nama_lengkap__icontains=search_query) |
            Q(mahasiswa__user__nrp__icontains=search_query) |
            Q(mahasiswa__user__email__icontains=search_query)
        )
    
    from django.db.models import Case, When, Value, IntegerField
    pendaftaran_list = pendaftaran_list.annotate(
        status_order=Case(
            When(status_pengajuan='pending', then=Value(1)),
            When(status_pengajuan='disetujui', then=Value(2)),
            When(status_pengajuan='ditolak', then=Value(3)),
            default=Value(4),
            output_field=IntegerField(),
        )
    ).order_by('status_order', '-created_at')
    
    total_pendaftaran = pendaftaran_list.count()
    menunggu_approval = pendaftaran_list.filter(status_pengajuan='pending').count()
    disetujui = pendaftaran_list.filter(status_pengajuan='disetujui').count()
    ditolak = pendaftaran_list.filter(status_pengajuan='ditolak').count()
    
    # DOWNLOAD
    if 'download' in request.GET:
        mahasiswa_id = request.GET.get('download')
        return download_foto_wajah(request, mahasiswa_id)
    
    # AJAX GET
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        action = request.GET.get('action', '')
        
        if action == 'detail':
            return get_detail_pendaftaran(request, request.GET.get('mahasiswa_id'))
        elif action == 'approve_modal':
            return get_approve_modal(request, request.GET.get('pengajuan_id'))
        elif action == 'reject_modal':
            return get_reject_modal(request, request.GET.get('pengajuan_id'))
    
    # ================= POST =================
    if request.method == 'POST':
        print("\n=== DEBUG: POST REQUEST DITERIMA ===")
        print(f"POST data: {dict(request.POST)}")
        
        pengajuan_id = request.POST.get('pengajuan_id')
        action = request.POST.get('action')
        alasan_penolakan = request.POST.get('alasan_penolakan', '')
        
        if not pengajuan_id:
            return JsonResponse({'success': False, 'message': 'ID tidak ditemukan'})
        
        try:
            pengajuan = Pengajuan_Pendaftaran.objects.get(id=pengajuan_id)
            mahasiswa = pengajuan.mahasiswa
            user = mahasiswa.user
            
            print(f"DEBUG: Status sebelum: {pengajuan.status_pengajuan}")
            print(f"DEBUG: Mahasiswa: {user.nama_lengkap}")
            
            with transaction.atomic():

                # ================= APPROVE =================
                if action == 'approve':
                    pengajuan.status_pengajuan = 'disetujui'
                    pengajuan.alasan_penolakan = ''
                    pengajuan.updated_at = datetime.now()
                    
                    user.is_active = True
                    user.status_akun = 'aktif'
                    
                    print("DEBUG: APPROVE dijalankan")
                    
                    pengajuan.save()
                    user.save()
                    
                    try:
                        subject = '✅ Pendaftaran Anda Telah Disetujui'
                        
                        context = {
                            'nama_mahasiswa': user.nama_lengkap,
                            'nim': user.nrp,
                            'status': 'disetujui',
                            'tgl_pengajuan': pengajuan.created_at.strftime('%d %B %Y'),
                            'tgl_keputusan': datetime.now().strftime('%d %B %Y'),
                            'pesan': 'Akun Anda telah diaktifkan.',
                        }
                        
                        html_message = render_to_string('email/notifikasi_pendaftaran.html', context)
                        plain_message = strip_tags(html_message)
                        
                        send_mail(
                            subject,
                            plain_message,
                            settings.DEFAULT_FROM_EMAIL,
                            [user.email],
                            html_message=html_message,
                            fail_silently=False,
                        )
                        
                        print(f"✅ Email approve terkirim ke {user.email}")
                    
                    except Exception as e:
                        print(f"❌ Email approve error: {e}")

                # ================= REJECT (FULL DEBUG) =================
                elif action == 'reject':
                    if not alasan_penolakan:
                        return JsonResponse({
                            'success': False,
                            'message': 'Alasan wajib diisi'
                        })
                    
                    pengajuan.status_pengajuan = 'ditolak'
                    pengajuan.alasan_penolakan = alasan_penolakan
                    pengajuan.updated_at = datetime.now()
                    
                    user.is_active = False
                    user.status_akun = 'nonaktif'
                    
                    print("DEBUG: REJECT dijalankan")
                    print(f"DEBUG: Alasan: {alasan_penolakan}")
                    
                    # SAVE DULU
                    pengajuan.save()
                    user.save()
                    
                    # ===== DEBUG EMAIL =====
                    print(f"\n📧 [EMAIL DEBUG] Kirim ke: {user.email}")
                    print(f"📧 Nama: {user.nama_lengkap}")
                    print(f"📧 NIM: {user.nrp}")
                    
                    # CEK TEMPLATE
                    from django.template.loader import get_template
                    try:
                        get_template('email/notifikasi_pendaftaran.html')
                        print("✅ Template ditemukan")
                    except Exception as e:
                        print(f"❌ Template error: {e}")
                    
                    try:
                        subject = '❌ Pendaftaran Anda Ditolak'
                        
                        context = {
                            'nama_mahasiswa': user.nama_lengkap,
                            'nim': user.nrp,
                            'status': 'ditolak',
                            'tgl_pengajuan': pengajuan.created_at.strftime('%d %B %Y'),
                            'tgl_keputusan': datetime.now().strftime('%d %B %Y'),
                            'alasan_penolakan': alasan_penolakan,
                            'pesan': 'Silakan daftar ulang.',
                        }
                        
                        print(f"📧 Context: {context}")
                        
                        html_message = render_to_string(
                            'email/notifikasi_pendaftaran.html', context
                        )
                        plain_message = strip_tags(html_message)
                        
                        print(f"📧 HTML length: {len(html_message)}")
                        
                        result = send_mail(
                            subject,
                            plain_message,
                            settings.DEFAULT_FROM_EMAIL,
                            [user.email],
                            html_message=html_message,
                            fail_silently=False,
                        )
                        
                        print(f"✅ Email reject terkirim ({result})")
                    
                    except Exception as e:
                        print(f"❌❌ EMAIL ERROR: {type(e).__name__}: {e}")
                        import traceback
                        traceback.print_exc()

                else:
                    return JsonResponse({'success': False, 'message': 'Action tidak valid'})
            
            return JsonResponse({
                'success': True,
                'message': 'Berhasil update status',
                'new_status': pengajuan.status_pengajuan
            })
        
        except Pengajuan_Pendaftaran.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Data tidak ditemukan'})
        
        except Exception as e:
            print(f"ERROR BESAR: {e}")
            import traceback
            traceback.print_exc()
            return JsonResponse({'success': False, 'message': str(e)})

    # ================= GET NORMAL =================
    return render(request, 'admin/approval_pendaftaran.html', {
        'pendaftaran_list': pendaftaran_list,
        'total_pendaftaran': total_pendaftaran,
        'menunggu_approval': menunggu_approval,
        'disetujui': disetujui,
        'ditolak': ditolak,
        'search_query': search_query,
    })

def get_detail_pendaftaran(request, mahasiswa_id):
    """Ambil data detail pendaftaran untuk modal"""
    mahasiswa = get_object_or_404(Mahasiswa, id=mahasiswa_id)
    
    # Ambil semua foto wajah (BUKAN KTM)
    foto_wajah = FotoWajah.objects.filter(mahasiswa=mahasiswa).order_by('-created_at')
    
    # Ambil foto KTM dari field mahasiswa
    foto_ktm = mahasiswa.foto_ktm  # ← Ambil dari model (field foto_ktm)
    
    # Ambil semua dosen pembimbing
    dosen_pembimbing_list = Mahasiswa_Dosen.objects.filter(
        mahasiswa=mahasiswa
    ).select_related('dosen').order_by('tipe_pembimbing')
    
    # Ambil kegiatan PA
    kegiatan_pa_list = mahasiswa.kegiatan_pa.all()
    
    context = {
        'mahasiswa': mahasiswa,
        'foto_wajah': foto_wajah,
        'foto_ktm': foto_ktm,  # ← KIRIM sebagai foto_ktm (bukan foto_ktp)
        'dosen_pembimbing_list': dosen_pembimbing_list,
        'kegiatan_pa_list': kegiatan_pa_list,
    }
    
    return render(request, 'admin/partials/detail_pendaftaran_content.html', context)

def get_approve_modal(request, pengajuan_id):
    """Ambil konten modal approve"""
    pengajuan = get_object_or_404(Pengajuan_Pendaftaran, id=pengajuan_id)
    
    context = {
        'pengajuan': pengajuan,
    }
    
    return render(request, 'admin/partials/approve_modal_content.html', context)

def get_reject_modal(request, pengajuan_id):
    """Ambil konten modal reject"""
    pengajuan = get_object_or_404(Pengajuan_Pendaftaran, id=pengajuan_id)
    
    context = {
        'pengajuan': pengajuan,
    }
    
    return render(request, 'admin/partials/reject_modal_content.html', context)

def download_foto_wajah(request, mahasiswa_id):
    """Download semua foto wajah mahasiswa dalam format zip"""
    mahasiswa = get_object_or_404(Mahasiswa, id=mahasiswa_id)
    foto_wajah_list = FotoWajah.objects.filter(mahasiswa=mahasiswa)
    
    if not foto_wajah_list:
        messages.error(request, "Tidak ada foto wajah untuk didownload")
        return redirect('approval_pendaftaran')
    
    # Create zip file in memory
    zip_buffer = io.BytesIO()
    
    try:
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for foto in foto_wajah_list:
                if foto.file_path and os.path.exists(foto.file_path.path):
                    file_name = f"{mahasiswa.user.nrp}_{mahasiswa.user.nama_lengkap}_{os.path.basename(foto.file_path.name)}"
                    zip_file.write(foto.file_path.path, file_name)
        
        zip_buffer.seek(0)
        
        # Create response
        response = HttpResponse(zip_buffer, content_type='application/zip')
        response['Content-Disposition'] = f'attachment; filename="foto_wajah_{mahasiswa.user.nrp}_{mahasiswa.user.nama_lengkap.replace(" ", "_")}.zip"'
        
        return response
    except Exception as e:
        print(f"Error creating zip: {e}")
        messages.error(request, "Gagal membuat file zip. Pastikan file foto tersedia.")
        return redirect('approval_pendaftaran')

def render_approval_page(request):
    """Render halaman utama approval"""
    pendaftaran_list = Pengajuan_Pendaftaran.objects.select_related(
        'mahasiswa__user', 
        'mahasiswa__jenjang_pendidikan',
        'mahasiswa__semester'
    ).all().order_by('-id')
    
    total_pendaftaran = pendaftaran_list.count()
    
    context = {
        'pendaftaran_list': pendaftaran_list,
        'total_pendaftaran': total_pendaftaran,
    }
    
    return render(request, 'admin/approval_pendaftaran.html', context)

@login_required
def data_mahasiswa(request):
    """View untuk data mahasiswa admin"""
    # Ambil parameter filter dari URL
    jenjang_filter = request.GET.get('jenjang', '')
    search_query = request.GET.get('search', '').strip()
    
    # Query dasar dengan semua relasi yang dibutuhkan
    mahasiswa_list = Mahasiswa.objects.filter(
        pengajuan_pendaftaran__status_pengajuan__iexact='disetujui'
    )
    
    # Filter berdasarkan jenjang pendidikan
    if jenjang_filter:
        mahasiswa_list = mahasiswa_list.filter(
            jenjang_pendidikan__nama_jenjang__iexact=jenjang_filter
        )
    
    # Filter berdasarkan search (nama, NRP, email)
    if search_query:
        mahasiswa_list = mahasiswa_list.filter(
            Q(user__nama_lengkap__icontains=search_query) |
            Q(user__nrp__icontains=search_query) |
            Q(user__email__icontains=search_query) |
            Q(jurusan__icontains=search_query) |
            Q(kelas__icontains=search_query)
        )
    
    # Urutkan berdasarkan NRP
    mahasiswa_list = mahasiswa_list.order_by('user__nrp')
    
    # Ambil semua jenjang pendidikan DENGAN ANNOTATE untuk hitung jumlah mahasiswa
    jenjang_pendidikan_list = Jenjang_Pendidikan.objects.annotate(
        jumlah_mahasiswa=Count('mahasiswa')
    )
    
    # Hitung total mahasiswa (setelah filter)
    total_mahasiswa = mahasiswa_list.count()
    
    # Handle export Excel
    if request.GET.get('export') == 'excel':
        return export_data_mahasiswa_excel(mahasiswa_list)
    
    context = {
        'mahasiswa_list': mahasiswa_list,
        'jenjang_pendidikan_list': jenjang_pendidikan_list,
        'total_mahasiswa': total_mahasiswa,
        'jenjang_filter': jenjang_filter,
        'search_query': search_query,
    }
    
    return render(request, 'admin/data_mahasiswa.html', context)

def export_data_mahasiswa_excel(mahasiswa_list):
    """Export data mahasiswa ke Excel"""
    # Buat workbook baru
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Mahasiswa"
    
    # Style untuk header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    alignment_center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Header kolom
    headers = [
        'No', 'NRP', 'Nama Mahasiswa', 'Email', 
        'Jenjang', 'Kelas', 'Semester', 'Jurusan/Prodi',
        'Status', 'Kegiatan SKS', 'Tanggal Daftar'
    ]
    
    # Tulis header
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment_center
        cell.border = thin_border
    
    # Tulis data
    for row_num, mahasiswa in enumerate(mahasiswa_list, 2):
        # Dapatkan kegiatan SKS sebagai string
        kegiatan_sks = ", ".join([k.nama_kegiatan for k in mahasiswa.kegiatan_pa.all()[:3]])
        if mahasiswa.kegiatan_pa.count() > 3:
            kegiatan_sks += f" (+{mahasiswa.kegiatan_pa.count() - 3} lagi)"
        
        data = [
            row_num - 1,  # No
            mahasiswa.user.nrp or "",
            mahasiswa.user.nama_lengkap or "",
            mahasiswa.user.email or "",
            mahasiswa.jenjang_pendidikan.nama_jenjang if mahasiswa.jenjang_pendidikan else "",
            mahasiswa.kelas or "",
            mahasiswa.semester.nama_semester if mahasiswa.semester else "",
            mahasiswa.jurusan or "",
            "Aktif" if mahasiswa.user.is_active else "Nonaktif",
            kegiatan_sks,
            mahasiswa.user.date_joined.strftime("%d-%m-%Y %H:%M") if mahasiswa.user.date_joined else ""
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            if col_num == 2:  # Kolom NRP - bold
                cell.font = Font(bold=True)
            if col_num in [1, 10]:  # Kolom No dan Status - center
                cell.alignment = alignment_center
    
    # Auto adjust column width
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column].width = adjusted_width
    
    # Buat response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"data_mahasiswa_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

@login_required
def edit_mahasiswa(request, mahasiswa_id):
    """View untuk edit data mahasiswa"""
    mahasiswa = get_object_or_404(Mahasiswa, id=mahasiswa_id)
    
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        # Ambil data yang dibutuhkan untuk form
        jenjang_pendidikan_list = Jenjang_Pendidikan.objects.all()
        semester_list = Semester.objects.all()
        kegiatan_pa_list = Kegiatan_PA.objects.all()
        
        context = {
            'mahasiswa': mahasiswa,
            'jenjang_pendidikan_list': jenjang_pendidikan_list,
            'semester_list': semester_list,
            'kegiatan_pa_list': kegiatan_pa_list,
        }
        
        return render(request, 'admin/partials/edit_mahasiswa_modal.html', context)
    
    # Handle POST request untuk update data
    if request.method == 'POST':
        try:
            # Update data user
            user = mahasiswa.user
            user.nama_lengkap = request.POST.get('nama_lengkap', user.nama_lengkap)
            user.nrp = request.POST.get('nrp', user.nrp)
            user.email = request.POST.get('email', user.email)
            
            # Update status aktif
            status_akun = request.POST.get('status_akun', 'aktif')
            user.is_active = (status_akun == 'aktif')
            user.save()
            
            # Update data mahasiswa
            jenjang_id = request.POST.get('jenjang_pendidikan')
            if jenjang_id:
                mahasiswa.jenjang_pendidikan = Jenjang_Pendidikan.objects.get(id=jenjang_id)
            
            semester_id = request.POST.get('semester')
            if semester_id:
                mahasiswa.semester = Semester.objects.get(id=semester_id)
            
            mahasiswa.kelas = request.POST.get('kelas', mahasiswa.kelas)
            mahasiswa.jurusan = request.POST.get('jurusan', mahasiswa.jurusan)
            mahasiswa.save()
            
            # Update kegiatan PA
            kegiatan_ids = request.POST.getlist('kegiatan_pa')
            mahasiswa.kegiatan_pa.set(Kegiatan_PA.objects.filter(id__in=kegiatan_ids))
            
            return JsonResponse({
                'success': True,
                'message': 'Data mahasiswa berhasil diperbarui!'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Gagal memperbarui data: {str(e)}'
            })
    
    return JsonResponse({
        'success': False,
        'message': 'Metode tidak diizinkan'
    }, status=405)

@login_required
def hapus_mahasiswa(request, mahasiswa_id):
    """View untuk hapus data mahasiswa"""
    if request.method == 'POST':
        mahasiswa = get_object_or_404(Mahasiswa, id=mahasiswa_id)
        
        try:
            # Hapus user terkait juga
            user = mahasiswa.user
            mahasiswa.delete()
            user.delete()
            
            return JsonResponse({
                'success': True,
                'message': 'Data mahasiswa berhasil dihapus'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Gagal menghapus data: {str(e)}'
            })
    
    return JsonResponse({
        'success': False,
        'message': 'Metode tidak diizinkan'
    }, status=405)

@login_required
def data_dosen(request):
    """View untuk data dosen admin"""
    if not (request.user.role == 'admin' or request.user.is_superuser):
        messages.error(request, 'Akses ditolak. Halaman ini untuk admin.')
        return redirect('admin_dashboard')
    
    # Ambil parameter filter dari URL
    prodi_filter = request.GET.get('prodi', '')
    search_query = request.GET.get('search', '').strip()
    
    # Query dasar
    dosen_list = Dosen.objects.all()
    
    # Filter berdasarkan prodi
    if prodi_filter:
        dosen_list = dosen_list.filter(prodi__icontains=prodi_filter)
    
    # Filter berdasarkan search (nama, NIP)
    if search_query:
        dosen_list = dosen_list.filter(
            Q(nama_dosen__icontains=search_query) |
            Q(nip__icontains=search_query) |
            Q(prodi__icontains=search_query)
        )
    
    # Urutkan berdasarkan nama
    dosen_list = dosen_list.order_by('nama_dosen')
    
    # Hitung statistik
    total_dosen = dosen_list.count()
    
    # Ambil semua prodi unik untuk filter
    prodi_list = Dosen.objects.values_list('prodi', flat=True).distinct().order_by('prodi')
    prodi_list = [p for p in prodi_list if p]  # Filter None/empty
    
    # Handle export Excel
    if request.GET.get('export') == 'excel':
        return export_data_dosen_excel(dosen_list)
    
    context = {
        'dosen_list': dosen_list,
        'total_dosen': total_dosen,
        'prodi_filter': prodi_filter,
        'search_query': search_query,
        'prodi_list': prodi_list,
    }
    
    return render(request, 'admin/data_dosen.html', context)

@login_required
def tambah_dosen(request):
    """View untuk menambah dosen baru"""
    if request.method == 'POST':
        try:
            nip = request.POST.get('nip', '').strip()
            nama_dosen = request.POST.get('nama_dosen', '').strip()
            prodi = request.POST.get('prodi', '').strip()
            
            if not nip or not nama_dosen or not prodi:
                return JsonResponse({'success': False, 'message': 'Semua field harus diisi'})
            
            # Cek duplikat NIP
            if Dosen.objects.filter(nip=nip).exists():
                return JsonResponse({'success': False, 'message': f'NIP {nip} sudah terdaftar'})
            
            dosen = Dosen.objects.create(
                nip=nip,
                nama_dosen=nama_dosen,
                prodi=prodi
            )
            
            return JsonResponse({'success': True, 'message': f'Dosen {dosen.nama_dosen} berhasil ditambahkan'})
            
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'})

@login_required
def edit_dosen(request, dosen_id):
    """View untuk mengedit dosen"""
    if request.method == 'POST':
        try:
            dosen = get_object_or_404(Dosen, id=dosen_id)
            
            nip = request.POST.get('nip', '').strip()
            nama_dosen = request.POST.get('nama_dosen', '').strip()
            prodi = request.POST.get('prodi', '').strip()
            
            if not nip or not nama_dosen or not prodi:
                return JsonResponse({'success': False, 'message': 'Semua field harus diisi'})
            
            # Cek duplikat NIP (kecuali dengan dirinya sendiri)
            if Dosen.objects.exclude(id=dosen_id).filter(nip=nip).exists():
                return JsonResponse({'success': False, 'message': f'NIP {nip} sudah digunakan dosen lain'})
            
            dosen.nip = nip
            dosen.nama_dosen = nama_dosen
            dosen.prodi = prodi
            dosen.save()
            
            return JsonResponse({'success': True, 'message': 'Data dosen berhasil diperbarui'})
            
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'})

@login_required
def hapus_dosen(request, dosen_id):
    """View untuk menghapus dosen"""
    if request.method == 'POST':
        try:
            dosen = get_object_or_404(Dosen, id=dosen_id)
            
            # Cek apakah dosen sedang digunakan sebagai pembimbing
            if Mahasiswa_Dosen.objects.filter(dosen=dosen).exists():
                return JsonResponse({
                    'success': False, 
                    'message': 'Tidak dapat menghapus dosen yang masih menjadi pembimbing mahasiswa'
                })
            
            nama_dosen = dosen.nama_dosen
            dosen.delete()
            
            return JsonResponse({'success': True, 'message': f'Dosen {nama_dosen} berhasil dihapus'})
            
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'})

@login_required
def get_detail_dosen(request, dosen_id):
    """API untuk mendapatkan detail dosen"""
    try:
        dosen = get_object_or_404(Dosen, id=dosen_id)
        
        # Hitung jumlah mahasiswa bimbingan
        jumlah_bimbingan = Mahasiswa_Dosen.objects.filter(dosen=dosen).count()
        
        return JsonResponse({
            'success': True,
            'dosen': {
                'id': dosen.id,
                'nip': dosen.nip,
                'nama_dosen': dosen.nama_dosen,
                'prodi': dosen.prodi,
                'jumlah_bimbingan': jumlah_bimbingan
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

def export_data_dosen_excel(dosen_list):
    """Export data dosen ke Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Dosen"
    
    # Style untuk header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    alignment_center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Header kolom
    headers = ['No', 'NIP', 'Nama Dosen', 'Program Studi', 'Jumlah Bimbingan']
    
    # Tulis header
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment_center
        cell.border = thin_border
    
    # Tulis data
    for row_num, dosen in enumerate(dosen_list, 2):
        jumlah_bimbingan = Mahasiswa_Dosen.objects.filter(dosen=dosen).count()
        
        data = [
            row_num - 1,
            dosen.nip,
            dosen.nama_dosen,
            dosen.prodi or '-',
            jumlah_bimbingan
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
    
    # Auto adjust column width
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column].width = adjusted_width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"data_dosen_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

# views.py - Tambahkan fungsi-fungsi berikut
@login_required
def master_kelas(request):
    """View untuk master data kelas admin"""
    if not (request.user.role == 'admin' or request.user.is_superuser):
        messages.error(request, 'Akses ditolak. Halaman ini untuk admin.')
        return redirect('admin_dashboard')
    
    # Ambil semua data kelas
    kelas_list = Kelas.objects.all().order_by('nama_kelas')
    
    # Filter berdasarkan pencarian
    search_query = request.GET.get('search', '')
    if search_query:
        kelas_list = kelas_list.filter(
            Q(nama_kelas__icontains=search_query) |
            Q(kode_kelas__icontains=search_query)
        )
    
    # Hitung statistik
    total_kelas = Kelas.objects.count()
    total_kelas_aktif = Kelas.objects.filter(is_active=True).count()
    total_mahasiswa = Mahasiswa.objects.filter(pengajuan_pendaftaran__status_pengajuan='disetujui').count()
    
    context = {
        'kelas_list': kelas_list,
        'total_kelas': total_kelas,
        'total_kelas_aktif': total_kelas_aktif,
        'total_mahasiswa': total_mahasiswa,
        'search_query': search_query,
    }
    
    return render(request, 'admin/master_kelas.html', context)


@login_required
def tambah_kelas(request):
    """View untuk menambah kelas baru"""
    if request.method == 'POST':
        try:
            nama_kelas = request.POST.get('nama_kelas', '').strip().upper()
            kode_kelas = request.POST.get('kode_kelas', '').strip().upper()
            is_active = request.POST.get('is_active') == 'on'
            
            if not nama_kelas:
                messages.error(request, 'Nama kelas harus diisi')
                return redirect('master_kelas')
            
            # Jika kode_kelas kosong, gunakan nama_kelas
            if not kode_kelas:
                kode_kelas = nama_kelas
            
            # Cek duplikat
            if Kelas.objects.filter(nama_kelas__iexact=nama_kelas).exists():
                messages.error(request, f'Kelas "{nama_kelas}" sudah ada')
                return redirect('master_kelas')
            
            if Kelas.objects.filter(kode_kelas__iexact=kode_kelas).exists():
                messages.error(request, f'Kode kelas "{kode_kelas}" sudah digunakan')
                return redirect('master_kelas')
            
            kelas = Kelas.objects.create(
                nama_kelas=nama_kelas,
                kode_kelas=kode_kelas,
                is_active=is_active
            )
            messages.success(request, f'Kelas "{kelas.nama_kelas}" berhasil ditambahkan')
            
        except Exception as e:
            messages.error(request, f'Terjadi kesalahan: {str(e)}')
    
    return redirect('master_kelas')


@login_required
def edit_kelas(request, kelas_id):
    """View untuk mengedit kelas"""
    if request.method == 'POST':
        try:
            kelas = get_object_or_404(Kelas, id=kelas_id)
            
            nama_kelas = request.POST.get('nama_kelas', '').strip().upper()
            kode_kelas = request.POST.get('kode_kelas', '').strip().upper()
            is_active = request.POST.get('is_active') == 'on'
            
            if not nama_kelas:
                return JsonResponse({'success': False, 'message': 'Nama kelas harus diisi'})
            
            # Jika kode_kelas kosong, gunakan nama_kelas
            if not kode_kelas:
                kode_kelas = nama_kelas
            
            # Cek duplikat (kecuali dengan dirinya sendiri)
            if Kelas.objects.exclude(id=kelas_id).filter(nama_kelas__iexact=nama_kelas).exists():
                return JsonResponse({'success': False, 'message': f'Kelas "{nama_kelas}" sudah ada'})
            
            if Kelas.objects.exclude(id=kelas_id).filter(kode_kelas__iexact=kode_kelas).exists():
                return JsonResponse({'success': False, 'message': f'Kode kelas "{kode_kelas}" sudah digunakan'})
            
            kelas.nama_kelas = nama_kelas
            kelas.kode_kelas = kode_kelas
            kelas.is_active = is_active
            kelas.save()
            
            return JsonResponse({'success': True, 'message': 'Kelas berhasil diperbarui'})
            
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'})


@login_required
def hapus_kelas(request, kelas_id):
    """View untuk menghapus kelas"""
    if request.method == 'POST':
        try:
            kelas = get_object_or_404(Kelas, id=kelas_id)
            
            # Cek apakah kelas sedang digunakan oleh mahasiswa
            if Mahasiswa.objects.filter(kelas=kelas).exists():
                return JsonResponse({
                    'success': False, 
                    'message': 'Tidak dapat menghapus kelas yang masih digunakan oleh mahasiswa'
                })
            
            nama_kelas = kelas.nama_kelas
            kelas.delete()
            
            return JsonResponse({'success': True, 'message': f'Kelas "{nama_kelas}" berhasil dihapus'})
            
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'})


@login_required
def get_detail_kelas(request, kelas_id):
    """API untuk mendapatkan detail kelas"""
    try:
        kelas = get_object_or_404(Kelas, id=kelas_id)
        
        # Hitung jumlah penggunaan
        mahasiswa_count = Mahasiswa.objects.filter(kelas=kelas).count()
        
        return JsonResponse({
            'success': True,
            'kelas': {
                'id': kelas.id,
                'nama_kelas': kelas.nama_kelas,
                'kode_kelas': kelas.kode_kelas,
                'is_active': kelas.is_active,
                'mahasiswa_count': mahasiswa_count
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@login_required
def master_prodi(request):
    """View untuk master data prodi/jurusan admin"""
    if not (request.user.role == 'admin' or request.user.is_superuser):
        messages.error(request, 'Akses ditolak. Halaman ini untuk admin.')
        return redirect('admin_dashboard')
    
    # Ambil parameter filter
    jenjang_filter = request.GET.get('jenjang', '')
    search_query = request.GET.get('search', '').strip()
    
    # Query data prodi
    prodi_list = Prodi.objects.select_related('jenjang').all()
    
    if jenjang_filter:
        prodi_list = prodi_list.filter(jenjang_id=jenjang_filter)
    
    if search_query:
        prodi_list = prodi_list.filter(
            Q(nama_prodi__icontains=search_query) |
            Q(kode_prodi__icontains=search_query)
        )
    
    prodi_list = prodi_list.order_by('jenjang__nama_jenjang', 'kode_prodi')
    
    # Statistik
    total_prodi = prodi_list.count()
    total_jenjang = Jenjang_Pendidikan.objects.count()
    total_mahasiswa = Mahasiswa.objects.filter(pengajuan_pendaftaran__status_pengajuan='disetujui').count()
    
    # Jenjang list untuk filter
    jenjang_list = Jenjang_Pendidikan.objects.all()
    
    context = {
        'prodi_list': prodi_list,
        'jenjang_list': jenjang_list,
        'total_prodi': total_prodi,
        'total_jenjang': total_jenjang,
        'total_mahasiswa': total_mahasiswa,
        'jenjang_filter': jenjang_filter,
        'search_query': search_query,
    }
    
    return render(request, 'admin/master_prodi.html', context)


@login_required
def tambah_prodi(request):
    """View untuk menambah prodi baru"""
    if request.method == 'POST':
        try:
            jenjang_id = request.POST.get('jenjang')
            kode_prodi = request.POST.get('kode_prodi', '').strip().upper()
            nama_prodi = request.POST.get('nama_prodi', '').strip()
            nama_singkat = request.POST.get('nama_singkat', '').strip()
            
            if not jenjang_id or not kode_prodi or not nama_prodi:
                return JsonResponse({'success': False, 'message': 'Semua field harus diisi'})
            
            # Cek duplikat
            if Prodi.objects.filter(kode_prodi=kode_prodi).exists():
                return JsonResponse({'success': False, 'message': f'Kode prodi {kode_prodi} sudah terdaftar'})
            
            prodi = Prodi.objects.create(
                jenjang_id=jenjang_id,
                kode_prodi=kode_prodi,
                nama_prodi=nama_prodi,
                nama_singkat=nama_singkat or kode_prodi,
                is_active=True
            )
            
            return JsonResponse({'success': True, 'message': f'Prodi {prodi.nama_prodi} berhasil ditambahkan'})
            
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'})


@login_required
def edit_prodi(request, prodi_id):
    """View untuk mengedit prodi"""
    if request.method == 'POST':
        try:
            prodi = get_object_or_404(Prodi, id=prodi_id)
            
            kode_prodi = request.POST.get('kode_prodi', '').strip().upper()
            nama_prodi = request.POST.get('nama_prodi', '').strip()
            nama_singkat = request.POST.get('nama_singkat', '').strip()
            is_active = request.POST.get('is_active') == 'on'
            
            if not kode_prodi or not nama_prodi:
                return JsonResponse({'success': False, 'message': 'Kode dan nama prodi harus diisi'})
            
            # Cek duplikat kode (kecuali dirinya sendiri)
            if Prodi.objects.exclude(id=prodi_id).filter(kode_prodi=kode_prodi).exists():
                return JsonResponse({'success': False, 'message': f'Kode prodi {kode_prodi} sudah digunakan prodi lain'})
            
            prodi.kode_prodi = kode_prodi
            prodi.nama_prodi = nama_prodi
            prodi.nama_singkat = nama_singkat or kode_prodi
            prodi.is_active = is_active
            prodi.save()
            
            return JsonResponse({'success': True, 'message': 'Prodi berhasil diperbarui'})
            
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'})


@login_required
def hapus_prodi(request, prodi_id):
    """View untuk menghapus prodi"""
    if request.method == 'POST':
        try:
            prodi = get_object_or_404(Prodi, id=prodi_id)
            
            # Cek apakah prodi digunakan mahasiswa
            if Mahasiswa.objects.filter(prodi=prodi).exists():
                return JsonResponse({
                    'success': False, 
                    'message': 'Tidak dapat menghapus prodi yang masih digunakan oleh mahasiswa'
                })
            
            nama_prodi = prodi.nama_prodi
            prodi.delete()
            
            return JsonResponse({'success': True, 'message': f'Prodi {nama_prodi} berhasil dihapus'})
            
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'})


@login_required
def get_detail_prodi(request, prodi_id):
    """API untuk mendapatkan detail prodi"""
    try:
        prodi = get_object_or_404(Prodi.objects.select_related('jenjang'), id=prodi_id)
        
        # Hitung jumlah penggunaan
        mahasiswa_count = Mahasiswa.objects.filter(prodi=prodi).count()
        
        return JsonResponse({
            'success': True,
            'prodi': {
                'id': prodi.id,
                'kode_prodi': prodi.kode_prodi,
                'nama_prodi': prodi.nama_prodi,
                'nama_singkat': prodi.nama_singkat,
                'jenjang_id': prodi.jenjang.id,
                'jenjang_nama': prodi.jenjang.nama_jenjang,
                'is_active': prodi.is_active,
                'mahasiswa_count': mahasiswa_count,
            }
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

def get_registration_summary(request):
    """API untuk mendapatkan ringkasan data registrasi"""
    try:
        step1_data = request.session.get('step1_data', {})
        step2_data = request.session.get('step2_data', {})
        
        return JsonResponse({
            'success': True,
            'step1': {
                'nama_lengkap': step1_data.get('nama_lengkap', ''),
                'nim': step1_data.get('nim', ''),
                'email': step1_data.get('email', ''),
                'kelas_nama': step1_data.get('kelas_nama', ''),
            },
            'step2': {
                'prodi_nama': step2_data.get('prodi_nama', ''),
                'jenjang_nama': step2_data.get('jenjang_nama', ''),
                'semester_nama': step2_data.get('semester_nama', ''),
                'kegiatan_pa_list': step2_data.get('kegiatan_pa_list', ''),
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@require_http_methods(["GET"])
@csrf_exempt
def get_prodi_by_jenjang(request, jenjang_id):
    try:
        prodi_list = Prodi.objects.filter(
            jenjang_id=jenjang_id,
            is_active=True
        ).values('id', 'nama_prodi', 'kode_prodi').order_by('nama_prodi')
        
        return JsonResponse({
            'success': True,
            'prodi_list': [
                {'id': p['id'], 'nama': p['nama_prodi'], 'kode': p['kode_prodi']} 
                for p in prodi_list
            ]
        })
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'prodi_list': [], 
            'error': str(e)
        })

@login_required
def master_data_wajah(request):
    """View untuk master data wajah admin"""
    # Ambil data mahasiswa dengan jumlah foto
    mahasiswa_list = Mahasiswa.objects.filter(
        pengajuan_pendaftaran__status_pengajuan='disetujui'
    ).select_related(
        'user', 'jenjang_pendidikan', 'semester'
    ).annotate(
        jumlah_foto=Count('foto_wajah')
    ).order_by('user__nama_lengkap')
    
    # Filter berdasarkan search
    search_query = request.GET.get('search', '')
    if search_query:
        mahasiswa_list = mahasiswa_list.filter(
            user__nama_lengkap__icontains=search_query
        ) | mahasiswa_list.filter(
            user__nrp__icontains=search_query
        ) | mahasiswa_list.filter(
            jurusan__icontains=search_query
        )
    
    context = {
        'mahasiswa_list': mahasiswa_list,
        'total_mahasiswa': Mahasiswa.objects.count(),
        'total_foto': FotoWajah.objects.count(),
        'search_query': search_query,
    }
    return render(request, 'admin/master_data_wajah.html', context)

@login_required
def get_foto_wajah_detail(request, mahasiswa_id):
    """API untuk mendapatkan detail foto wajah mahasiswa (JSON response)"""
    mahasiswa = get_object_or_404(
        Mahasiswa.objects.select_related('user', 'jenjang_pendidikan', 'semester'),
        id=mahasiswa_id
    )
    
    foto_list = FotoWajah.objects.filter(mahasiswa=mahasiswa).order_by('-created_at')
    
    # Siapkan data untuk response JSON
    foto_data = []
    for foto in foto_list:
        foto_data.append({
            'id': foto.id,
            'url': foto.file_path.url,
            'created_at': foto.created_at.strftime("%d %b %Y %H:%M"),
            'keterangan': foto.keterangan or ""
        })
    
    response_data = {
        'success': True,
        'mahasiswa': {
            'id': mahasiswa.id,
            'nama_lengkap': mahasiswa.user.nama_lengkap,
            'nrp': mahasiswa.user.nrp,
            'jurusan': mahasiswa.jurusan,
            'jenjang': mahasiswa.jenjang_pendidikan.nama_jenjang if mahasiswa.jenjang_pendidikan else '-',
            'semester': mahasiswa.semester.nama_semester if mahasiswa.semester else '-',
            'status_akun': mahasiswa.user.get_status_akun_display(),
            'is_active': mahasiswa.user.status_akun == 'aktif',
            'date_joined': mahasiswa.user.date_joined.strftime("%d %b %Y"),
        },
        'fotos': foto_data,
        'total_fotos': len(foto_data)
    }
    
    return JsonResponse(response_data)

@login_required
def download_all_fotos(request, mahasiswa_id):
    """Download semua foto mahasiswa dalam format zip"""
    mahasiswa = get_object_or_404(Mahasiswa, id=mahasiswa_id)
    foto_list = FotoWajah.objects.filter(mahasiswa=mahasiswa)
    
    if not foto_list:
        return JsonResponse({'error': 'Tidak ada foto untuk didownload'}, status=404)
    
    # Create zip file in memory
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for foto in foto_list:
            if foto.file_path:
                try:
                    # Baca file dari storage
                    file_content = foto.file_path.read()
                    # Buat nama file yang rapi
                    filename = f"{mahasiswa.user.nrp}_{foto.id:03d}_{foto.created_at.strftime('%Y%m%d_%H%M%S')}.jpg"
                    zip_file.writestr(filename, file_content)
                except Exception as e:
                    continue
    
    zip_buffer.seek(0)
    
    response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="foto_wajah_{mahasiswa.user.nrp}.zip"'
    return response

@login_required
def data_sks(request):
    """View untuk data SKS/Kegiatan admin"""
    # Ambil semua kegiatan
    kegiatan_list = Kegiatan_PA.objects.select_related('jenjang_pendidikan', 'tahun_ajaran').all().order_by('jenjang_pendidikan__nama_jenjang', 'nama_kegiatan')
    
    # Ambil semua jenjang untuk filter
    jenjang_list = Jenjang_Pendidikan.objects.all()
    
    # Ambil semua tahun ajaran untuk tab semester
    tahun_ajaran_list = Tahun_Ajaran.objects.all().order_by('-tanggal_mulai')
    
    # Filter berdasarkan jenjang
    jenjang_filter = request.GET.get('jenjang', '')
    if jenjang_filter:
        kegiatan_list = kegiatan_list.filter(jenjang_pendidikan_id=jenjang_filter)
    
    # Filter berdasarkan pencarian
    search_query = request.GET.get('search', '')
    if search_query:
        kegiatan_list = kegiatan_list.filter(
            Q(nama_kegiatan__icontains=search_query) |
            Q(jenjang_pendidikan__nama_jenjang__icontains=search_query)
        )
    
    context = {
        'kegiatan_list': kegiatan_list,
        'jenjang_list': jenjang_list,
        'tahun_ajaran_list': tahun_ajaran_list,
        'total_kegiatan': kegiatan_list.count(),
        'jenjang_filter': jenjang_filter,
        'search_query': search_query,
    }
    return render(request, 'admin/data_sks.html', context)

@login_required
def master_jenjang(request):
    """View untuk master data jenjang pendidikan"""
    if not (request.user.role == 'admin' or request.user.is_superuser):
        messages.error(request, 'Akses ditolak. Halaman ini untuk admin.')
        return redirect('admin_dashboard')
    
    # Ambil semua data jenjang
    jenjang_list = Jenjang_Pendidikan.objects.all().order_by('nama_jenjang')
    
    # Filter berdasarkan pencarian
    search_query = request.GET.get('search', '')
    if search_query:
        jenjang_list = jenjang_list.filter(nama_jenjang__icontains=search_query)
    
    # Hitung statistik
    total_jenjang = jenjang_list.count()
    total_mahasiswa = Mahasiswa.objects.filter(jenjang_pendidikan__in=jenjang_list).count()
    total_kegiatan = Kegiatan_PA.objects.filter(jenjang_pendidikan__in=jenjang_list).count()
    
    context = {
        'jenjang_list': jenjang_list,
        'total_jenjang': total_jenjang,
        'total_mahasiswa': total_mahasiswa,
        'total_kegiatan': total_kegiatan,
        'search_query': search_query,
    }
    
    return render(request, 'admin/master_jenjang.html', context)

@login_required
def tambah_jenjang(request):
    """View untuk menambah jenjang pendidikan"""
    if request.method == 'POST':
        try:
            nama_jenjang = request.POST.get('nama_jenjang', '').strip()
            
            if not nama_jenjang:
                messages.error(request, 'Nama jenjang harus diisi')
                return redirect('master_jenjang')
            
            # Cek duplikat
            if Jenjang_Pendidikan.objects.filter(nama_jenjang__iexact=nama_jenjang).exists():
                messages.error(request, f'Jenjang "{nama_jenjang}" sudah ada')
                return redirect('master_jenjang')
            
            jenjang = Jenjang_Pendidikan.objects.create(nama_jenjang=nama_jenjang)
            messages.success(request, f'Jenjang "{jenjang.nama_jenjang}" berhasil ditambahkan')
            
        except Exception as e:
            messages.error(request, f'Terjadi kesalahan: {str(e)}')
    
    return redirect('master_jenjang')

@login_required
def edit_jenjang(request, jenjang_id):
    """View untuk mengedit jenjang pendidikan"""
    if request.method == 'POST':
        try:
            jenjang = get_object_or_404(Jenjang_Pendidikan, id=jenjang_id)
            nama_jenjang_baru = request.POST.get('nama_jenjang', '').strip()
            
            if not nama_jenjang_baru:
                return JsonResponse({'success': False, 'message': 'Nama jenjang harus diisi'})
            
            # Cek duplikat (kecuali dengan dirinya sendiri)
            if Jenjang_Pendidikan.objects.exclude(id=jenjang_id).filter(nama_jenjang__iexact=nama_jenjang_baru).exists():
                return JsonResponse({'success': False, 'message': f'Jenjang "{nama_jenjang_baru}" sudah ada'})
            
            jenjang.nama_jenjang = nama_jenjang_baru
            jenjang.save()
            
            return JsonResponse({'success': True, 'message': 'Jenjang berhasil diperbarui'})
            
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'})

@login_required
def hapus_jenjang(request, jenjang_id):
    """View untuk menghapus jenjang pendidikan"""
    if request.method == 'POST':
        try:
            jenjang = get_object_or_404(Jenjang_Pendidikan, id=jenjang_id)
            
            # Cek apakah jenjang sedang digunakan
            if Mahasiswa.objects.filter(jenjang_pendidikan=jenjang).exists():
                return JsonResponse({
                    'success': False, 
                    'message': 'Tidak dapat menghapus jenjang yang masih digunakan oleh mahasiswa'
                })
            
            if Kegiatan_PA.objects.filter(jenjang_pendidikan=jenjang).exists():
                return JsonResponse({
                    'success': False, 
                    'message': 'Tidak dapat menghapus jenjang yang masih digunakan oleh kegiatan SKS'
                })
            
            nama_jenjang = jenjang.nama_jenjang
            jenjang.delete()
            
            return JsonResponse({'success': True, 'message': f'Jenjang "{nama_jenjang}" berhasil dihapus'})
            
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'})

@login_required
def get_detail_jenjang(request, jenjang_id):
    """API untuk mendapatkan detail jenjang"""
    try:
        jenjang = get_object_or_404(Jenjang_Pendidikan, id=jenjang_id)
        
        # Hitung jumlah penggunaan
        mahasiswa_count = Mahasiswa.objects.filter(jenjang_pendidikan=jenjang).count()
        kegiatan_count = Kegiatan_PA.objects.filter(jenjang_pendidikan=jenjang).count()
        
        return JsonResponse({
            'success': True,
            'jenjang': {
                'id': jenjang.id,
                'nama_jenjang': jenjang.nama_jenjang,
                'mahasiswa_count': mahasiswa_count,
                'kegiatan_count': kegiatan_count
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })
        
@login_required
def tambah_kegiatan_sks(request):
    """View untuk menambah kegiatan SKS baru"""
    if request.method == 'POST':
        try:
            # Ambil data dari form
            nama_kegiatan = request.POST.get('nama_kegiatan')
            jenjang_id = request.POST.get('jenjang_pendidikan')
            tahun_ajaran_id = request.POST.get('tahun_ajaran')
            jumlah_sks = request.POST.get('jumlah_sks', 0)
            total_jam_minggu = request.POST.get('total_jam_minggu', 0)
            target_jam = request.POST.get('target_jam', 0)
            
            # Validasi data
            if not nama_kegiatan or not jenjang_id:
                messages.error(request, 'Nama kegiatan dan jenjang harus diisi')
                return redirect('data_sks')
            
            # Jika tahun ajaran tidak dipilih, ambil tahun ajaran aktif
            if not tahun_ajaran_id:
                tahun_ajaran_aktif = Tahun_Ajaran.objects.filter(status_aktif='aktif').first()
                if tahun_ajaran_aktif:
                    tahun_ajaran_id = tahun_ajaran_aktif.id
                else:
                    messages.error(request, 'Tidak ada tahun ajaran aktif. Silakan aktifkan tahun ajaran terlebih dahulu.')
                    return redirect('data_sks')
            
            # Buat objek kegiatan baru
            kegiatan = Kegiatan_PA(
                nama_kegiatan=nama_kegiatan,
                jenjang_pendidikan_id=jenjang_id,
                tahun_ajaran_id=tahun_ajaran_id,
                jumlah_sks=jumlah_sks,
                total_jam_minggu=total_jam_minggu,
                target_jam=target_jam
            )
            kegiatan.save()
            
            messages.success(request, 'Kegiatan SKS berhasil ditambahkan')
            return redirect('data_sks')
            
        except Exception as e:
            messages.error(request, f'Terjadi kesalahan: {str(e)}')
            return redirect('data_sks')
    
    return redirect('data_sks')

@login_required
def edit_kegiatan_sks(request, kegiatan_id):
    """View untuk mengedit kegiatan SKS"""
    if request.method == 'POST':
        try:
            kegiatan = get_object_or_404(Kegiatan_PA, id=kegiatan_id)
            
            # Update data
            kegiatan.nama_kegiatan = request.POST.get('nama_kegiatan', kegiatan.nama_kegiatan)
            kegiatan.jenjang_pendidikan_id = request.POST.get('jenjang_pendidikan', kegiatan.jenjang_pendidikan_id)
            kegiatan.tahun_ajaran_id = request.POST.get('tahun_ajaran', kegiatan.tahun_ajaran_id)
            kegiatan.jumlah_sks = request.POST.get('jumlah_sks', kegiatan.jumlah_sks)
            kegiatan.total_jam_minggu = request.POST.get('total_jam_minggu', kegiatan.total_jam_minggu)
            kegiatan.target_jam = request.POST.get('target_jam', kegiatan.target_jam)
            
            kegiatan.save()
            messages.success(request, 'Kegiatan SKS berhasil diperbarui')
            
        except Exception as e:
            messages.error(request, f'Terjadi kesalahan: {str(e)}')
    
    return redirect('data_sks')

@login_required
def hapus_kegiatan_sks(request, kegiatan_id):
    """View untuk menghapus kegiatan SKS"""
    if request.method == 'POST':
        try:
            kegiatan = get_object_or_404(Kegiatan_PA, id=kegiatan_id)
            
            # Cek apakah kegiatan sudah digunakan
            from .models import Status_Pemenuhan_SKS
            if Status_Pemenuhan_SKS.objects.filter(kegiatan_pa=kegiatan).exists():
                messages.error(request, 'Tidak dapat menghapus kegiatan yang sudah digunakan')
                return redirect('data_sks')
            
            kegiatan.delete()
            messages.success(request, 'Kegiatan SKS berhasil dihapus')
            
        except Exception as e:
            messages.error(request, f'Terjadi kesalahan: {str(e)}')
    
    return redirect('data_sks')

@login_required
def get_detail_kegiatan(request, kegiatan_id):
    """API untuk mendapatkan detail kegiatan (JSON response)"""
    try:
        kegiatan = get_object_or_404(
            Kegiatan_PA.objects.select_related('jenjang_pendidikan', 'tahun_ajaran'),
            id=kegiatan_id
        )
        
        response_data = {
            'success': True,
            'kegiatan': {
                'id': kegiatan.id,
                'nama_kegiatan': kegiatan.nama_kegiatan,
                'jenjang_pendidikan': {
                    'id': kegiatan.jenjang_pendidikan.id,
                    'nama': kegiatan.jenjang_pendidikan.nama_jenjang
                },
                'tahun_ajaran': {
                    'id': kegiatan.tahun_ajaran.id if kegiatan.tahun_ajaran else None,
                    'nama': kegiatan.tahun_ajaran.nama_tahun_ajaran if kegiatan.tahun_ajaran else None
                },
                'jumlah_sks': kegiatan.jumlah_sks,
                'total_jam_minggu': kegiatan.total_jam_minggu,
                'target_jam': kegiatan.target_jam
            }
        }
        
        return JsonResponse(response_data)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)

@login_required
def tambah_tahun_ajaran(request):
    """View untuk menambah tahun ajaran baru"""
    if request.method == 'POST':
        try:
            nama_tahun_ajaran = request.POST.get('nama_tahun_ajaran')
            tanggal_mulai = request.POST.get('tanggal_mulai')
            tanggal_selesai = request.POST.get('tanggal_selesai')
            
            if not nama_tahun_ajaran or not tanggal_mulai or not tanggal_selesai:
                messages.error(request, 'Semua field harus diisi')
                return redirect('data_sks')
            
            tahun_ajaran = Tahun_Ajaran(
                nama_tahun_ajaran=nama_tahun_ajaran,
                tanggal_mulai=tanggal_mulai,
                tanggal_selesai=tanggal_selesai,
                status_aktif='nonaktif'
            )
            tahun_ajaran.save()
            
            messages.success(request, 'Tahun ajaran berhasil ditambahkan')
            return redirect('data_sks')
            
        except Exception as e:
            messages.error(request, f'Terjadi kesalahan: {str(e)}')
    
    return redirect('data_sks')

@login_required
def edit_tahun_ajaran(request, tahun_id):
    """View untuk mengedit tahun ajaran"""
    if request.method == 'POST':
        try:
            tahun_ajaran = get_object_or_404(Tahun_Ajaran, id=tahun_id)
            
            tahun_ajaran.nama_tahun_ajaran = request.POST.get('nama_tahun_ajaran', tahun_ajaran.nama_tahun_ajaran)
            tahun_ajaran.tanggal_mulai = request.POST.get('tanggal_mulai', tahun_ajaran.tanggal_mulai)
            tahun_ajaran.tanggal_selesai = request.POST.get('tanggal_selesai', tahun_ajaran.tanggal_selesai)
            
            tahun_ajaran.save()
            messages.success(request, 'Tahun ajaran berhasil diperbarui')
            
        except Exception as e:
            messages.error(request, f'Terjadi kesalahan: {str(e)}')
    
    return redirect('data_sks')

@login_required
def hapus_tahun_ajaran(request, tahun_id):
    """View untuk menghapus tahun ajaran"""
    if request.method == 'POST':
        try:
            tahun_ajaran = get_object_or_404(Tahun_Ajaran, id=tahun_id)
            
            # Cek apakah tahun ajaran digunakan di kegiatan
            if Kegiatan_PA.objects.filter(tahun_ajaran=tahun_ajaran).exists():
                messages.error(request, 'Tidak dapat menghapus tahun ajaran yang sudah digunakan')
                return redirect('data_sks')
            
            tahun_ajaran.delete()
            messages.success(request, 'Tahun ajaran berhasil dihapus')
            
        except Exception as e:
            messages.error(request, f'Terjadi kesalahan: {str(e)}')
    
    return redirect('data_sks')

@login_required
def get_detail_tahun_ajaran(request, tahun_id):
    """API untuk mendapatkan detail tahun ajaran"""
    try:
        tahun_ajaran = get_object_or_404(Tahun_Ajaran, id=tahun_id)
        
        response_data = {
            'success': True,
            'tahun_ajaran': {
                'id': tahun_ajaran.id,
                'nama_tahun_ajaran': tahun_ajaran.nama_tahun_ajaran,
                'tanggal_mulai': tahun_ajaran.tanggal_mulai.strftime('%Y-%m-%d'),
                'tanggal_selesai': tahun_ajaran.tanggal_selesai.strftime('%Y-%m-%d'),
                'status_aktif': tahun_ajaran.status_aktif
            }
        }
        
        return JsonResponse(response_data)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)

@login_required
def aktifkan_tahun_ajaran(request):
    """View untuk mengaktifkan tahun ajaran"""
    if request.method == 'POST':
        try:
            tahun_id = request.POST.get('tahun_id')
            
            # Nonaktifkan semua tahun ajaran terlebih dahulu
            Tahun_Ajaran.objects.update(status_aktif='nonaktif')
            
            # Aktifkan tahun ajaran yang dipilih
            tahun_ajaran = get_object_or_404(Tahun_Ajaran, id=tahun_id)
            tahun_ajaran.status_aktif = 'aktif'
            tahun_ajaran.save()
            
            return JsonResponse({'success': True})
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=400)
    
    return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)

@login_required
def rekap_presensi(request):
    """View untuk rekap presensi admin dengan filter"""
    
    if not (request.user.role == 'admin' or request.user.is_superuser):
        messages.error(request, 'Akses ditolak. Halaman untuk admin saja.')
        return redirect('login')
    
    try:
        tahun_ajaran_aktif = Tahun_Ajaran.objects.get(status_aktif='aktif')
        semester_aktif_info = f"{tahun_ajaran_aktif.nama_tahun_ajaran} (Aktif)"
    except Tahun_Ajaran.DoesNotExist:
        semester_aktif_info = "Tidak ada Tahun Ajaran Aktif"
    
    # 2. Gunakan form yang sudah ada di forms.py
    form = FilterRekapPresensiForm(request.GET or None)
    
    # 3. Query awal dengan semua relasi yang dibutuhkan
    presensi_list = Presensi.objects.select_related(
        'mahasiswa__user',
        'mahasiswa__jenjang_pendidikan',
        'kegiatan_pa'  # PASTIKAN INI DIPILIH
    ).all().order_by('-tanggal_presensi', '-jam_checkin')
    
    # 4. Terapkan filter jika form valid
    if form.is_valid():
        tanggal_mulai = form.cleaned_data.get('tanggal_mulai')
        tanggal_selesai = form.cleaned_data.get('tanggal_selesai')
        tingkatan = form.cleaned_data.get('tingkatan')
        kegiatan = form.cleaned_data.get('kegiatan')
        
        # Filter tanggal
        if tanggal_mulai:
            presensi_list = presensi_list.filter(tanggal_presensi__gte=tanggal_mulai)
        if tanggal_selesai:
            presensi_list = presensi_list.filter(tanggal_presensi__lte=tanggal_selesai)
        
        # Filter tingkatan - PERBAIKAN INI
        if tingkatan:  
        # Jika menggunakan ModelChoiceField, tingkatan adalah objek
            presensi_list = presensi_list.filter(
                mahasiswa__jenjang_pendidikan=tingkatan  # Langsung bandingkan dengan objek
            )
        
        # Filter kegiatan - gunakan many-to-many pada mahasiswa (karena sistem agregat)
        if kegiatan:
            presensi_list = presensi_list.filter(mahasiswa__kegiatan_pa=kegiatan)

    # 5. Default: 7 hari terakhir jika tidak ada filter tanggal
    if not request.GET.get('tanggal_mulai') and not request.GET.get('tanggal_selesai'):
        default_end = datetime.now().date()
        default_start = default_end - timedelta(days=7)
        presensi_list = presensi_list.filter(
            tanggal_presensi__gte=default_start,
            tanggal_presensi__lte=default_end
        )
        # Set initial form values
        form.initial = {
            'tanggal_mulai': default_start,
            'tanggal_selesai': default_end
        }
    
    # 6. Siapkan data untuk template
    data_presensi = []
    
    # Prefetch untuk optimasi ManyToMany
    presensi_list = presensi_list.prefetch_related('mahasiswa__kegiatan_pa')
    
    for presensi in presensi_list:
        # Hitung durasi
        durasi_str = "-"
        if presensi.jam_checkin and presensi.jam_checkout:
            try:
                checkin_dt = datetime.combine(presensi.tanggal_presensi, presensi.jam_checkin)
                checkout_dt = datetime.combine(presensi.tanggal_presensi, presensi.jam_checkout)
                
                if checkout_dt < checkin_dt:
                    checkout_dt += timedelta(days=1)
                
                delta = checkout_dt - checkin_dt
                total_seconds = delta.total_seconds()
                
                hours = int(total_seconds // 3600)
                minutes = int((total_seconds % 3600) // 60)
                
                if hours > 0:
                    durasi_str = f"{hours}j {minutes}m" if minutes > 0 else f"{hours}j"
                elif minutes > 0:
                    durasi_str = f"{minutes}m"
                else:
                    durasi_str = "0m"
                    
            except Exception as e:
                durasi_str = "Error"
        
        # Ambil tingkatan dari jenjang pendidikan mahasiswa
        tingkatan_value = "-"
        if presensi.mahasiswa.jenjang_pendidikan:
            tingkatan_value = presensi.mahasiswa.jenjang_pendidikan.nama_jenjang
        
        # Ambil SEMUA kegiatan PA yang diambil mahasiswa (sesuai permintaan user)
        kegiatan_pa_list = presensi.mahasiswa.kegiatan_pa.all()
        if kegiatan_pa_list:
            kegiatan_pa_value = ", ".join([k.nama_kegiatan for k in kegiatan_pa_list])
        else:
            kegiatan_pa_value = "-"
        
        data_presensi.append({
            'tanggal': presensi.tanggal_presensi,
            'nrp': presensi.mahasiswa.nim,
            'nama': presensi.mahasiswa.user.nama_lengkap,
            'tingkatan': tingkatan_value,  # Sekarang harus muncul
            'kegiatan_pa': kegiatan_pa_value,  # Sekarang hanya 1 kegiatan sesuai presensi
            'check_in': presensi.jam_checkin,
            'check_out': presensi.jam_checkout,
            'durasi': durasi_str,
        })
    
    # 7. Konteks untuk template
    context = {
        'form': form,
        'data_presensi': data_presensi,
        'semester': semester_aktif_info,
        'total_presensi': len(data_presensi),
    }
    
    return render(request, 'admin/rekap_presensi.html', context)

def _save_base64_to_file(base64_str, prefix):
    """Helper function to save base64 image to ContentFile"""
    try:
        # Handle both formats: with and without data:image prefix
        if ';base64,' in base64_str:
            format, imgstr = base64_str.split(';base64,')
            ext = format.split('/')[-1]
        else:
            imgstr = base64_str
            ext = 'jpg'  # default extension
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_%f')[:-3]
        filename = f'{prefix}_{timestamp}.{ext}'
        return ContentFile(base64.b64decode(imgstr), name=filename)
    except Exception as e:
        print(f"[ERROR] _save_base64_to_file: {e}")
        raise